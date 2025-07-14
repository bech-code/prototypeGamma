from django.shortcuts import render
from rest_framework import viewsets, permissions
import rest_framework.status as rf_status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.utils import timezone
from django.db.models import Q, Count, F, Avg, Sum
from django.core.paginator import Paginator
from .utils import calculate_distance
import requests
import json
import logging
from decimal import Decimal
import math
from datetime import timedelta
# Temporairement commenté pour éviter l'erreur GDAL
# from django.contrib.gis.geos import Point
# from django.contrib.gis.db.models.functions import Distance
from django.core.mail import send_mail
# Ajout import Twilio
try:
    from twilio.rest import Client as TwilioClient
except ImportError:
    TwilioClient = None
import os
from django.contrib.auth.models import Permission, Group
from django.contrib.auth import get_user_model
from .serializers import (
    ClientSerializer,
    TechnicianSerializer,
    RepairRequestSerializer,
    RequestDocumentSerializer,
    ReviewSerializer,
    PaymentSerializer,
    ConversationSerializer,
    MessageSerializer,
    MessageAttachmentSerializer,
    NotificationSerializer,
    TechnicianLocationSerializer,
    SystemConfigurationSerializer,
    CinetPayPaymentSerializer,
    CinetPayInitiationSerializer,
    CinetPayNotificationSerializer,
    RepairRequestCreateSerializer,
    TechnicianNearbySerializer,
    PermissionSerializer,
    GroupSerializer,
    PlatformConfigurationSerializer,
    ClientLocationSerializer,
    ReportSerializer,
    AdminNotificationSerializer,
    AuditLogSerializer,
    SubscriptionPaymentRequestSerializer,
    TechnicianSubscriptionSerializer,
    ChatConversationSerializer,
    ChatMessageSerializer,
    ChatMessageAttachmentSerializer,
)
from .models import (
    Client, Technician, RepairRequest, RequestDocument, Review, Payment, Conversation, Message, Notification, MessageAttachment, TechnicianLocation, SystemConfiguration, CinetPayPayment, PlatformConfiguration, ClientLocation,
    Report, AdminNotification, SubscriptionPaymentRequest, TechnicianSubscription,
    ChatConversation, ChatMessage, ChatMessageAttachment,
)
from rest_framework.views import APIView
from users.models import AuditLog
from django.utils.timezone import now
from rest_framework.pagination import PageNumberPagination
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils.decorators import classonlymethod
from .cinetpay import init_cinetpay_payment
from django.http import HttpResponse
import csv
import openpyxl

logger = logging.getLogger(__name__)

# ============================================================================
# FONCTIONS UTILITAIRES
# ============================================================================

def get_technician_profile(user):
    """
    Récupère le profil technicien d'un utilisateur en essayant les deux relations possibles.
    Retourne le premier profil trouvé ou None si aucun n'existe.
    """
    import logging
    logger = logging.getLogger(__name__)
    # Essayer d'abord technician_depannage (relation dans l'app depannage)
    technician = getattr(user, 'technician_depannage', None)
    if technician:
        logger.info(f"[DEBUG] get_technician_profile: technician_depannage trouvé: {technician} (type: {type(technician)})")
        return technician
    # Essayer ensuite technician_profile (relation dans l'app users)
    technician = getattr(user, 'technician_profile', None)
    if technician:
        logger.info(f"[DEBUG] get_technician_profile: technician_profile trouvé: {technician} (type: {type(technician)})")
        return technician
    logger.info(f"[DEBUG] get_technician_profile: Aucun profil technicien trouvé pour user {user}")
    return None

# ============================================================================
# VUES DE TEST PUBLIQUES
# ============================================================================


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def find_nearest_technician(request):
    """Find the nearest available technician based on user's location."""
    try:
        user_latitude = float(request.data.get("latitude"))
        user_longitude = float(request.data.get("longitude"))
    except (TypeError, ValueError):
        return Response(
            {"error": "Latitude et longitude invalides"},
            status=400,
        )

    # Récupérer tous les techniciens disponibles
    available_technicians = Technician.objects.filter(is_available=True)

    if not available_technicians:
        return Response(
            {"error": "Aucun technicien disponible pour le moment"},
            status=404,
        )

    # Calculer la distance pour chaque technicien
    technicians_with_distance = []
    for technician in available_technicians:
        distance = calculate_distance(
            user_latitude, user_longitude, technician.current_latitude, technician.current_longitude
        )
        technicians_with_distance.append((technician, distance))

    # Trier par distance et prendre le plus proche
    nearest_technician, distance = min(technicians_with_distance, key=lambda x: x[1])

    # Sérialiser le technicien et ajouter la distance
    serializer = TechnicianSerializer(nearest_technician)
    response_data = serializer.data
    response_data["distance"] = round(distance, 2)

    return Response(response_data)


class PublicTestViewSet(viewsets.ViewSet):
    """ViewSet pour les tests publics de l'API."""

    permission_classes = [AllowAny]

    @action(detail=False, methods=["get"])
    def health_check(self, request):
        """Vérification de santé de l'API."""
        return Response(
            {
                "status": "healthy",
                "message": "API DepanneTeliman fonctionne correctement",
                "timestamp": timezone.now().isoformat(),
                "version": "1.0.0",
            }
        )

    @action(detail=False, methods=["get"])
    def api_info(self, request):
        """Informations sur l'API."""
        return Response(
            {
                "name": "DepanneTeliman API",
                "description": "API pour la gestion des services de dépannage",
                "endpoints": {
                    "auth": "/users/login/",
                    "register": "/users/register/",
                    "repair_requests": "/depannage/api/repair-requests/",
                    "technicians": "/depannage/api/technicians/",
                    "clients": "/depannage/api/clients/",
                    "admin": "/admin/",
                },
                "authentication": "JWT (Bearer Token)",
                "documentation": "Consultez les endpoints pour plus d'informations",
            }
        )


# ============================================================================
# VIEWSETS EXISTANTES
# ============================================================================


class CsrfExemptMixin(object):
    @classonlymethod
    def as_view(cls, *args, **kwargs):
        view = super().as_view(*args, **kwargs)
        return csrf_exempt(view)

class RepairRequestViewSet(CsrfExemptMixin, viewsets.ModelViewSet):
    """ViewSet pour gérer les demandes de réparation."""

    serializer_class = RepairRequestSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_serializer_class(self):
        if self.action == "create":
            return RepairRequestCreateSerializer
        return RepairRequestSerializer

    def get_queryset(self):
        """Optimise les requêtes avec select_related et prefetch_related."""
        user = self.request.user
        
        # Base queryset avec optimisations
        queryset = RepairRequest.objects.select_related(
            'client__user',
            'technician__user'
        ).prefetch_related(
            'documents',
            'review',
            'payments',
            'notifications'
        )
        
        # Filtrage selon le type d'utilisateur
        if user.is_superuser or user.user_type == 'admin':
            # Admin voit toutes les demandes
            return queryset
        
        elif user.user_type == 'technician':
            # Technicien voit ses demandes assignées et les demandes disponibles
            technician = get_technician_profile(user)
            if technician:
                return queryset.filter(
                    Q(technician=technician) | 
                    Q(status='pending', specialty_needed=technician.specialty)
                )
            return RepairRequest.objects.none()
        
        else:
            # Client voit ses propres demandes
            try:
                client = user.client_profile
                return queryset.filter(client=client)
            except:
                return RepairRequest.objects.none()

    def perform_create(self, serializer):
        """Crée une demande avec validation et notifications."""
        user = self.request.user
        data = serializer.validated_data
        # Vérification de la géolocalisation
        if not data.get('latitude') or not data.get('longitude'):
            raise ValidationError("La géolocalisation est obligatoire")
        # Vérification du profil client
        if not hasattr(user, 'client_profile') or user.client_profile is None:
            from rest_framework.response import Response
            from rest_framework import status
            raise ValidationError("Seuls les clients avec un profil complet peuvent créer une demande de réparation.")
        repair_request = serializer.save(client=user.client_profile)
        self.notify_available_technicians(repair_request)
        logger.info(f"Demande créée: {repair_request.id} par {user.username}")
        return repair_request

    def notify_available_technicians(self, repair_request):
        """Notifie les techniciens disponibles de la nouvelle demande avec algorithme de matching intelligent."""
        try:
            from .models import calculate_distance
            
            # Trouver les techniciens disponibles dans la spécialité avec critères optimisés
            available_technicians = Technician.objects.filter(
                is_available=True,
                specialty=repair_request.specialty_needed,
                is_verified=True,
                current_latitude__isnull=False,
                current_longitude__isnull=False
            ).select_related('user').prefetch_related('subscriptions')
            
            # Filtrer les techniciens avec abonnement actif
            technicians_with_subscription = []
            for tech in available_technicians:
                if tech.has_active_subscription:
                    technicians_with_subscription.append(tech)
            
            if not technicians_with_subscription:
                logger.warning(f"Aucun technicien disponible pour la spécialité {repair_request.specialty_needed}")
                return
            
            # Calculer les distances et scores pour chaque technicien
            tech_with_scores = []
            for technician in technicians_with_subscription:
                # Calculer la distance
                distance = calculate_distance(
                    repair_request.latitude, repair_request.longitude,
                    technician.current_latitude, technician.current_longitude
                )
                
                # Vérifier si le technicien est dans son rayon de service
                if distance > technician.service_radius_km:
                    continue
                
                # Calculer le score de matching
                score = self.calculate_technician_score(technician, repair_request, distance)
                
                tech_with_scores.append({
                    'technician': technician,
                    'distance': distance,
                    'score': score
                })
            
            # Trier par score décroissant (meilleurs techniciens en premier)
            tech_with_scores.sort(key=lambda x: x['score'], reverse=True)
            
            # Prendre les 10 meilleurs techniciens pour la notification
            top_technicians = tech_with_scores[:10]
            
            # Créer des notifications optimisées
            notifications = []
            for tech_data in top_technicians:
                technician = tech_data['technician']
                distance = tech_data['distance']
                score = tech_data['score']
                
                # Message personnalisé selon l'urgence
                urgency_text = ""
                if repair_request.urgency_level == 'sos':
                    urgency_text = "🚨 DEMANDE SOS - Intervention immédiate requise"
                elif repair_request.urgency_level == 'urgent':
                    urgency_text = "⚡ DEMANDE URGENTE - Intervention rapide"
                else:
                    urgency_text = "📋 Nouvelle demande disponible"
                
                notification = Notification(
                    recipient=technician.user,
                    type=Notification.Type.URGENT_REQUEST,
                    title=f"{urgency_text}",
                    message=f"{repair_request.title} - {distance:.1f}km - {repair_request.address}",
                    request=repair_request,
                    extra_data={
                        'request_id': repair_request.id,
                        'specialty': repair_request.specialty_needed,
                        'urgency': repair_request.urgency_level,
                        'distance': distance,
                        'score': score,
                        'estimated_price': repair_request.estimated_price,
                        'client_name': f"{repair_request.client.user.first_name} {repair_request.client.user.last_name}",
                        'client_phone': repair_request.client.phone
                    }
                )
                notifications.append(notification)
            
            # Bulk create pour optimiser les performances
            if notifications:
                Notification.objects.bulk_create(notifications)
                logger.info(f"Notifications envoyées à {len(notifications)} techniciens pour la demande {repair_request.id}")
                
                # Log détaillé pour debugging
                for i, tech_data in enumerate(top_technicians[:3]):  # Top 3 seulement
                    tech = tech_data['technician']
                    logger.info(f"Top {i+1}: {tech.user.username} - Distance: {tech_data['distance']:.1f}km - Score: {tech_data['score']:.2f}")
                
        except Exception as e:
            logger.error(f"Erreur lors de la notification des techniciens: {e}")

    def calculate_technician_score(self, technician, repair_request, distance):
        """Calcule un score de matching pour un technicien basé sur plusieurs critères."""
        score = 0
        
        # Score de distance (40% du score total)
        max_distance = technician.service_radius_km
        distance_score = max(0, (max_distance - distance) / max_distance) * 40
        score += distance_score
        
        # Score de disponibilité (20% du score total)
        # Vérifier le nombre de demandes en cours
        active_requests = RepairRequest.objects.filter(
            technician=technician,
            status__in=['assigned', 'in_progress']
        ).count()
        
        availability_score = max(0, (5 - active_requests) / 5) * 20
        score += availability_score
        
        # Score de réputation (20% du score total)
        rating_score = (technician.average_rating / 5) * 20
        score += rating_score
        
        # Score d'urgence (20% du score total)
        urgency_multiplier = {
            'sos': 1.5,
            'urgent': 1.3,
            'same_day': 1.1,
            'normal': 1.0
        }
        urgency_score = 20 * urgency_multiplier.get(repair_request.urgency_level, 1.0)
        score += urgency_score
        
        # Bonus pour les techniciens premium
        if technician.has_premium_subscription:
            score += 10
        
        return score

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def assign_technician(self, request, pk=None):
        """Assigne un technicien à une demande avec validation."""
        repair_request = self.get_object()
        user = request.user
        
        # Validation des permissions
        if user.user_type not in ['technician', 'admin']:
            return Response(
                {"error": "Seuls les techniciens peuvent s'assigner aux demandes"},
                status=403
            )
        
        try:
            technician = get_technician_profile(user)
            if not technician:
                return Response(
                    {"error": "Profil technicien non trouvé"},
                    status=404
                )
            
            # Vérifications de sécurité
            if repair_request.status != RepairRequest.Status.PENDING:
                return Response(
                    {"error": "Cette demande n'est plus disponible"},
                    status=400
                )
            
            if repair_request.specialty_needed != technician.specialty:
                return Response(
                    {"error": "Spécialité non compatible"},
                    status=400
                )
            
            # Assignation du technicien
            repair_request.assign_to_technician(technician)
            
            # Notification au client
            Notification.objects.create(
                recipient=repair_request.client.user,
                type=Notification.Type.REQUEST_ASSIGNED,
                title="Technicien assigné",
                message=f"Un technicien a accepté votre demande: {repair_request.title}",
                request=repair_request
            )
            
            # Notification au technicien
            Notification.objects.create(
                recipient=technician.user,
                type=Notification.Type.REQUEST_ASSIGNED,
                title="Demande acceptée",
                message=f"Vous avez accepté la demande: {repair_request.title}",
                request=repair_request
            )
            
            serializer = self.get_serializer(repair_request)
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Erreur lors de l'assignation: {e}")
            return Response(
                {"error": "Erreur lors de l'assignation"},
                status=500
            )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def update_status(self, request, pk=None):
        """Met à jour le statut d'une demande avec validation."""
        repair_request = self.get_object()
        user = request.user
        new_status = request.data.get('status')
        
        # Validation du statut
        valid_statuses = dict(RepairRequest.Status.choices)
        if new_status not in valid_statuses:
            return Response(
                {"error": "Statut invalide"},
                status=400
            )
        
        # Vérification des permissions
        if user.user_type == 'technician':
            technician = get_technician_profile(user)
            if not technician or repair_request.technician != technician:
                return Response(
                    {"error": "Vous ne pouvez modifier que vos propres demandes"},
                    status=403
                )
        
        try:
            old_status = repair_request.status
            repair_request.status = new_status
            
            # Actions spécifiques selon le statut
            if new_status == RepairRequest.Status.IN_PROGRESS:
                repair_request.start_work()
            elif new_status == RepairRequest.Status.COMPLETED:
                final_price = request.data.get('final_price')
                repair_request.complete_work(final_price)
            
            repair_request.save()
            
            # Notification du changement de statut
            self.notify_status_change(repair_request, old_status, new_status)
            
            serializer = self.get_serializer(repair_request)
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour du statut: {e}")
            return Response(
                {"error": "Erreur lors de la mise à jour"},
                status=500
            )

    def notify_status_change(self, repair_request, old_status, new_status):
        """Notifie le changement de statut aux parties concernées."""
        status_messages = {
            RepairRequest.Status.IN_PROGRESS: "Le travail a commencé",
            RepairRequest.Status.COMPLETED: "Le travail est terminé",
            RepairRequest.Status.CANCELLED: "La demande a été annulée"
        }
        
        message = status_messages.get(new_status, f"Statut changé vers {new_status}")
        
        # Notification au client
        if repair_request.client:
            Notification.objects.create(
                recipient=repair_request.client.user,
                type=Notification.Type.REQUEST_STARTED if new_status == RepairRequest.Status.IN_PROGRESS else Notification.Type.REQUEST_COMPLETED,
                title=f"Demande {new_status}",
                message=f"{message}: {repair_request.title}",
                request=repair_request
            )
        
        # Notification au technicien
        if repair_request.technician:
            Notification.objects.create(
                recipient=repair_request.technician.user,
                type=Notification.Type.REQUEST_STARTED if new_status == RepairRequest.Status.IN_PROGRESS else Notification.Type.REQUEST_COMPLETED,
                title=f"Demande {new_status}",
                message=f"{message}: {repair_request.title}",
                request=repair_request
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def dashboard_stats(self, request):
        """Statistiques optimisées pour le dashboard."""
        user = request.user
        
        try:
            if user.user_type == 'technician':
                technician = get_technician_profile(user)
                if not technician:
                    return Response({"error": "Profil technicien non trouvé"}, status=404)
                
                # Requêtes optimisées avec annotations
                stats = RepairRequest.objects.filter(
                    technician=technician
                ).aggregate(
                    total_requests=Count('id'),
                    completed_requests=Count('id', filter=Q(status='completed')),
                    pending_requests=Count('id', filter=Q(status='pending')),
                    in_progress_requests=Count('id', filter=Q(status='in_progress')),
                    total_earnings=Sum('final_price', filter=Q(status='completed')),
                    avg_rating=Avg('review__rating', filter=Q(status='completed', review__isnull=False))
                )
                
                # Calcul du taux de réussite
                total = stats['total_requests'] or 0
                completed = stats['completed_requests'] or 0
                success_rate = round((completed / total * 100) if total > 0 else 0, 1)
                
                return Response({
                    'total_requests': total,
                    'completed_requests': completed,
                    'pending_requests': stats['pending_requests'] or 0,
                    'in_progress_requests': stats['in_progress_requests'] or 0,
                    'total_earnings': float(stats['total_earnings'] or 0),
                    'average_rating': round(stats['avg_rating'] or 0, 1),
                    'success_rate': success_rate,
                    'is_available': technician.is_available,
                    'specialty': technician.specialty,
                    'experience_level': technician.experience_level
                })
            
            elif user.user_type == 'client':
                client = user.client_profile
                
                stats = RepairRequest.objects.filter(
                    client=client
                ).aggregate(
                    total_requests=Count('id'),
                    completed_requests=Count('id', filter=Q(status='completed')),
                    pending_requests=Count('id', filter=Q(status='pending')),
                    in_progress_requests=Count('id', filter=Q(status='in_progress')),
                    total_spent=Sum('final_price', filter=Q(status='completed'))
                )
                
                return Response({
                    'total_requests': stats['total_requests'] or 0,
                    'completed_requests': stats['completed_requests'] or 0,
                    'pending_requests': stats['pending_requests'] or 0,
                    'in_progress_requests': stats['in_progress_requests'] or 0,
                    'total_spent': float(stats['total_spent'] or 0)
                })
            
            else:
                # Admin - statistiques globales
                stats = RepairRequest.objects.aggregate(
                    total_requests=Count('id'),
                    completed_requests=Count('id', filter=Q(status='completed')),
                    pending_requests=Count('id', filter=Q(status='pending')),
                    total_revenue=Sum('final_price', filter=Q(status='completed'))
                )
                
                return Response({
                    'total_requests': stats['total_requests'] or 0,
                    'completed_requests': stats['completed_requests'] or 0,
                    'pending_requests': stats['pending_requests'] or 0,
                    'total_revenue': float(stats['total_revenue'] or 0)
                })
                
        except Exception as e:
            logger.error(f"Erreur lors du calcul des statistiques: {e}")
            return Response(
                {"error": "Erreur lors du calcul des statistiques"},
                status=500
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def available_technicians(self, request):
        """Récupère les techniciens disponibles avec géolocalisation optimisée."""
        try:
            # Paramètres de géolocalisation
            latitude = request.query_params.get('latitude')
            longitude = request.query_params.get('longitude')
            specialty = request.query_params.get('specialty')
            max_distance = float(request.query_params.get('max_distance', 20))  # km
            
            # Base queryset optimisée
            queryset = Technician.objects.filter(
                is_available=True,
                is_verified=True
            ).select_related('user')
            
            # Filtrage par spécialité si spécifiée
            if specialty:
                queryset = queryset.filter(specialty=specialty)
            
            # Calcul des distances si géolocalisation fournie
            if latitude and longitude:
                try:
                    user_lat = float(latitude)
                    user_lon = float(longitude)
                    
                    technicians_with_distance = []
                    for technician in queryset:
                        if technician.current_latitude and technician.current_longitude:
                            distance = calculate_distance(
                                user_lat, user_lon,
                                technician.current_latitude, technician.current_longitude
                            )
                            if distance <= max_distance:
                                technicians_with_distance.append((technician, distance))
                    
                    # Tri par distance
                    technicians_with_distance.sort(key=lambda x: x[1])
                    technicians = [tech for tech, _ in technicians_with_distance]
                    
                except (ValueError, TypeError):
                    return Response(
                        {"error": "Coordonnées géographiques invalides"},
                        status=400
                    )
            else:
                # Sans géolocalisation, retourner tous les techniciens disponibles
                technicians = list(queryset)
            
            # Pagination
            paginator = PageNumberPagination()
            paginator.page_size = 20
            paginated_technicians = paginator.paginate_queryset(technicians, request)
            
            serializer = TechnicianSerializer(paginated_technicians, many=True)
            return paginator.get_paginated_response(serializer.data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des techniciens: {e}")
            return Response(
                {"error": "Erreur lors de la récupération des techniciens"},
                status=500
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def project_statistics(self, request):
        """Récupère les statistiques complètes du projet (Admin seulement)."""
        user = request.user
        
        if user.user_type != "admin":
            return Response(
                {"error": "Accès non autorisé"}, status=403
            )

        from django.utils import timezone
        from django.db.models import Count, Sum, Avg, Q
        from datetime import timedelta
        from users.models import User, AuditLog
        from decimal import Decimal

        now = timezone.now()
        last_30_days = now - timedelta(days=30)
        last_7_days = now - timedelta(days=7)
        last_24_hours = now - timedelta(hours=24)

        # Statistiques utilisateurs
        total_users = User.objects.count()
        total_clients = User.objects.filter(user_type='client').count()
        total_technicians = User.objects.filter(user_type='technician').count()
        total_admins = User.objects.filter(user_type='admin').count()
        
        # Utilisateurs actifs (avec activité récente) - version robuste
        active_client_ids = Client.objects.filter(
            repair_requests__created_at__gte=last_30_days
        ).values_list('user_id', flat=True)
        active_technician_ids = Technician.objects.filter(
            repair_requests__created_at__gte=last_30_days
        ).values_list('user_id', flat=True)
        active_users_30d = User.objects.filter(
            Q(id__in=active_client_ids) | Q(id__in=active_technician_ids)
        ).distinct().count()

        # Statistiques demandes
        total_requests = RepairRequest.objects.count()
        pending_requests = RepairRequest.objects.filter(status="pending").count()
        in_progress_requests = RepairRequest.objects.filter(status="in_progress").count()
        completed_requests = RepairRequest.objects.filter(status="completed").count()
        cancelled_requests = RepairRequest.objects.filter(status="cancelled").count()
        
        # Demandes récentes
        requests_24h = RepairRequest.objects.filter(created_at__gte=last_24_hours).count()
        requests_7d = RepairRequest.objects.filter(created_at__gte=last_7_days).count()
        requests_30d = RepairRequest.objects.filter(created_at__gte=last_30_days).count()

        # Statistiques financières
        total_revenue = Payment.objects.filter(
            status='completed', 
            payment_type='client_payment'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        total_payouts = Payment.objects.filter(
            status='completed', 
            payment_type='technician_payout'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        platform_fees = total_revenue - total_payouts

        # Statistiques par spécialité
        specialty_stats = (
            RepairRequest.objects.values("specialty_needed")
            .annotate(
                count=Count("id"),
                completed=Count("id", filter=Q(status="completed")),
                avg_price=Avg("final_price")
            )
            .order_by("-count")
        )

        # Statistiques techniciens
        total_verified_technicians = Technician.objects.filter(is_verified=True).count()
        total_available_technicians = Technician.objects.filter(is_available=True, is_verified=True).count()
        
        # Top techniciens par performance
        top_technicians = (
            Technician.objects.filter(is_verified=True)
            .annotate(
                total_jobs=Count("repair_requests", filter=Q(repair_requests__status="completed")),
                avg_rating=Avg("repair_requests__review__rating", filter=Q(repair_requests__status="completed")),
                total_earnings=Sum("repair_requests__payments__amount", filter=Q(repair_requests__payments__status="completed"))
            )
            .filter(total_jobs__gt=0)
            .order_by("-total_jobs")[:10]
        )

        # Statistiques de satisfaction
        total_reviews = Review.objects.count()
        avg_rating = Review.objects.aggregate(avg=Avg('rating'))['avg'] or 0
        satisfaction_rate = (
            Review.objects.filter(rating__gte=4).count() / total_reviews * 100
        ) if total_reviews > 0 else 0

        # Statistiques de sécurité
        total_logins = AuditLog.objects.filter(event_type='login', status='success').count()
        failed_logins = AuditLog.objects.filter(event_type='login', status='failure').count()
        security_alerts = AuditLog.objects.filter(risk_score__gte=80).count()
        
        # Statistiques temporelles (évolution)
        daily_requests = []
        for i in range(7):
            date = now - timedelta(days=i)
            count = RepairRequest.objects.filter(
                created_at__date=date.date()
            ).count()
            daily_requests.append({
                'date': date.strftime('%Y-%m-%d'),
                'count': count
            })
        daily_requests.reverse()

        # Statistiques géographiques
        city_stats = (
            RepairRequest.objects.values("city")
            .annotate(count=Count("id"))
            .filter(city__isnull=False)
            .exclude(city="")
            .order_by("-count")[:10]
        )

        # Statistiques de paiement
        payment_methods = (
            Payment.objects.values("method")
            .annotate(count=Count("id"), total=Sum("amount"))
            .filter(status="completed")
            .order_by("-total")
        )

        # Préparer les données de réponse
        response_data = {
            # Vue d'ensemble
            'overview': {
                'total_users': total_users,
                'total_clients': total_clients,
                'total_technicians': total_technicians,
                'total_admins': total_admins,
                'active_users_30d': active_users_30d,
                'total_requests': total_requests,
                'completed_requests': completed_requests,
                'total_revenue': float(total_revenue),
                'platform_fees': float(platform_fees),
                'avg_rating': round(avg_rating, 1),
                'satisfaction_rate': round(satisfaction_rate, 1)
            },
            
            # Demandes
            'requests': {
                'total': total_requests,
                'pending': pending_requests,
                'in_progress': in_progress_requests,
                'completed': completed_requests,
                'cancelled': cancelled_requests,
                'recent_24h': requests_24h,
                'recent_7d': requests_7d,
                'recent_30d': requests_30d,
                'daily_evolution': daily_requests
            },
            
            # Financier
            'financial': {
                'total_revenue': float(total_revenue),
                'total_payouts': float(total_payouts),
                'platform_fees': float(platform_fees),
                'payment_methods': list(payment_methods)
            },
            
            # Spécialités
            'specialties': {
                'stats': list(specialty_stats),
                'top_technicians': [
                    {
                        'id': tech.id,
                        'name': tech.user.get_full_name() or tech.user.username,
                        'specialty': tech.get_specialty_display(),
                        'total_jobs': tech.total_jobs,
                        'avg_rating': round(tech.avg_rating or 0, 1),
                        'total_earnings': float(tech.total_earnings or 0)
                    }
                    for tech in top_technicians
                ]
            },
            
            # Techniciens
            'technicians': {
                'total': total_technicians,
                'verified': total_verified_technicians,
                'available': total_available_technicians,
                'availability_rate': round((total_available_technicians / total_verified_technicians * 100) if total_verified_technicians > 0 else 0, 1)
            },
            
            # Satisfaction
            'satisfaction': {
                'total_reviews': total_reviews,
                'avg_rating': round(avg_rating, 1),
                'satisfaction_rate': round(satisfaction_rate, 1)
            },
            
            # Sécurité
            'security': {
                'total_logins': total_logins,
                'failed_logins': failed_logins,
                'security_alerts': security_alerts,
                'success_rate': round((total_logins / (total_logins + failed_logins) * 100) if (total_logins + failed_logins) > 0 else 0, 1)
            },
            
            # Géographie
            'geography': {
                'top_cities': list(city_stats)
            }
        }

        # DEBUG : Affiche la réponse dans la console
        print("=== project_statistics response ===")
        import pprint; pprint.pprint(response_data)
        return Response(response_data)

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def notification_candidates(self, request):
        """Endpoint temporaire pour diagnostiquer les techniciens candidats à la notification."""
        specialty = request.query_params.get("specialty")
        lat = request.query_params.get("lat")
        lng = request.query_params.get("lng")
        if not specialty or not lat or not lng:
            return Response({"error": "specialty, lat et lng requis"}, status=400)
        try:
            lat = float(lat)
            lng = float(lng)
        except ValueError:
            return Response({"error": "lat/lng invalides"}, status=400)
        technicians = Technician.objects.filter(
            specialty=specialty,
            is_available=True,
            is_verified=True,
            current_latitude__isnull=False,
            current_longitude__isnull=False,
        )
        tech_with_distance = []
        for tech in technicians:
            distance = calculate_distance(lat, lng, tech.current_latitude, tech.current_longitude)
            tech_with_distance.append({
                "tech_id": tech.id,
                "user_id": tech.user.id,
                "username": tech.user.username,
                "specialty": tech.specialty,
                "distance_km": round(distance, 2),
                "is_available": tech.is_available,
                "is_verified": tech.is_verified,
                "lat": tech.current_latitude,
                "lng": tech.current_longitude
            })
        tech_with_distance.sort(key=lambda x: x["distance_km"])
        return Response({"candidates": tech_with_distance})

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def report_no_show(self, request, pk=None):
        """Signaler que le technicien n'est pas venu : notification admin + réassignation automatique (max 2 fois)."""
        repair_request = self.get_object()
        user = request.user
        from users.models import User
        # Incrémenter le compteur d'absence
        repair_request.no_show_count = (repair_request.no_show_count or 0) + 1
        repair_request.save()
        # Notifier l'admin
        admin_users = User.objects.filter(is_staff=True, is_active=True)
        for admin in admin_users:
            Notification.objects.create(
                recipient=admin,
                title="Technicien non venu",
                message=f"Le client a signalé l'absence du technicien pour la demande #{repair_request.id}",
                type="technician_no_show",
                request=repair_request,
            )
        # Limite de réaffectation automatique
        if repair_request.no_show_count > 2:
            for admin in admin_users:
                Notification.objects.create(
                    recipient=admin,
                    title="Intervention humaine requise",
                    message=f"La demande #{repair_request.id} a déjà été réaffectée 2 fois. Intervention manuelle nécessaire.",
                    type="manual_reassignment_required",
                    request=repair_request,
                )
            Notification.objects.create(
                recipient=repair_request.client.user,
                title="Intervention admin requise",
                message=f"Nous n'avons pas pu réassigner automatiquement un technicien pour votre demande #{repair_request.id}. Un admin va vous contacter.",
                type="manual_reassignment_required",
                request=repair_request,
            )
            return Response({"success": False, "message": "Limite de réaffectation automatique atteinte. Intervention admin requise."}, status=400)
        # Exclure le technicien précédent
        previous_technician = repair_request.technician
        # Relancer le matching (mêmes critères, hors technicien précédent)
        lat = repair_request.latitude
        lng = repair_request.longitude
        technicians = Technician.objects.filter(
            specialty=repair_request.specialty_needed,
            is_available=True,
            is_verified=True,
            current_latitude__isnull=False,
            current_longitude__isnull=False,
        ).exclude(id=previous_technician.id if previous_technician else None)
        technicians = [t for t in technicians if t.has_active_subscription]
        busy_tech_ids = set(
            RepairRequest.objects.filter(
                status__in=[RepairRequest.Status.ASSIGNED, RepairRequest.Status.IN_PROGRESS]
            ).values_list('technician_id', flat=True)
        )
        technicians = [t for t in technicians if t.id not in busy_tech_ids]
        MIN_RATING = 3.5
        technicians = [t for t in technicians if t.average_rating >= MIN_RATING]
        tech_with_distance = []
        for tech in technicians:
            distance = calculate_distance(lat, lng, tech.current_latitude, tech.current_longitude)
            if distance <= tech.service_radius_km:
                tech_with_distance.append((tech, distance))
        tech_with_distance.sort(key=lambda x: x[1])
        closest_techs = [t[0] for t in tech_with_distance[:10]]
        # Réassigner au premier dispo
        if closest_techs:
            new_technician = closest_techs[0]
            repair_request.technician = new_technician
            repair_request.status = RepairRequest.Status.ASSIGNED
            repair_request.save()
            # Notifier le nouveau technicien
            Notification.objects.create(
                recipient=new_technician.user,
                title="Nouvelle demande réassignée",
                message=f"Vous avez été réassigné à la demande #{repair_request.id}",
                type="new_request_technician",
                request=repair_request,
            )
            # Notifier le client avec le nom du nouveau technicien
            Notification.objects.create(
                recipient=repair_request.client.user,
                title="Nouveau technicien en route",
                message=f"{new_technician.user.get_full_name() or new_technician.user.username} a été réassigné à votre demande #{repair_request.id}",
                type="technician_assigned",
                request=repair_request,
            )
            # Retirer l'ancien technicien des participants si différent du nouveau
            old_technician = None
            if repair_request.conversation.participants.count() > 1:
                for participant in repair_request.conversation.participants.all():
                    if hasattr(participant, 'technician_depannage') and participant != new_technician.user:
                        old_technician = participant
                        break
            if old_technician:
                repair_request.conversation.participants.remove(old_technician)
            # Ajouter le nouveau technicien
            repair_request.conversation.participants.add(new_technician.user)

            # Ajouter le client aux participants
            repair_request.conversation.participants.add(repair_request.client.user)

            return Response({"success": True, "message": f"Demande réassignée à {new_technician.user.get_full_name() or new_technician.user.username}."})
        else:
            # Aucun technicien dispo
            for admin in admin_users:
                Notification.objects.create(
                    recipient=admin,
                    title="Aucun technicien disponible",
                    message=f"Aucun technicien n'a pu être réassigné à la demande #{repair_request.id}",
                    type="no_technician_available",
                    request=repair_request,
                )
            Notification.objects.create(
                recipient=repair_request.client.user,
                title="Aucun technicien disponible",
                message=f"Nous n'avons pas pu réassigner un technicien pour votre demande #{repair_request.id}. Un admin va vous contacter.",
                type="no_technician_available",
                request=repair_request,
            )
            return Response({"success": False, "message": "Aucun technicien disponible pour la réaffectation."}, status=400)

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def validate_mission(self, request, pk=None):
        """Valide la mission côté client, envoie le reçu et retourne les infos du reçu."""
        repair_request = self.get_object()
        user = request.user
        if not hasattr(user, 'client_profile') or repair_request.client.user != user:
            return Response({"error": "Seul le client peut valider la mission."}, status=403)
        if not repair_request.status == RepairRequest.Status.COMPLETED:
            return Response({"error": "La mission doit être terminée pour être validée."}, status=400)
        if repair_request.mission_validated:
            return Response({"error": "La mission a déjà été validée."}, status=400)
        if repair_request.status in [RepairRequest.Status.CANCELLED, RepairRequest.Status.CANCELLED]:
            return Response({"error": "Impossible de valider une mission annulée ou refusée."}, status=400)
        # Marquer comme validée
        repair_request.mission_validated = True
        repair_request.save()
        # Envoi du reçu par email
        try:
            subject = "Reçu de mission - Merci pour votre confiance !"
            message = (
                f"Bonjour {user.get_full_name()},\n\n"
                f"Votre mission #{repair_request.id} a été validée le {repair_request.completed_at.strftime('%d/%m/%Y à %Hh%M')}.\n"
                f"Technicien : {repair_request.technician.user.get_full_name()}\n"
                f"Service : {repair_request.title}\n"
                f"Adresse : {repair_request.address}\n"
                f"Référence : {repair_request.uuid}\n"
                f"Paiement : effectué en main propre au technicien.\n\n"
                f"Merci d'avoir choisi notre plateforme !"
            )
            send_mail(
                subject,
                message,
                'no-reply@votreservice.ml',
                [user.email],
                fail_silently=True,
            )
        except Exception as e:
            logger.error(f"Erreur lors de l'envoi du reçu de mission : {e}")
        # Notification pour le reçu
        Notification.objects.create(
            recipient=user,
            title="Reçu de mission",
            message=f"Votre reçu de mission #{repair_request.id} est disponible.",
            type="system",
            request=repair_request,
        )
        
        # Notification pour encourager la notation
        Notification.objects.create(
            recipient=user,
            title="Partagez votre expérience !",
            message=f"Votre mission avec {repair_request.technician.user.get_full_name() or repair_request.technician.user.username} est terminée. Aidez d'autres clients en notant votre technicien !",
            type="review_reminder",
            request=repair_request
        )
        # Retourner les infos du reçu
        data = {
            "success": True,
            "date": repair_request.completed_at,
            "technician": repair_request.technician.user.get_full_name() if repair_request.technician else None,
            "service": repair_request.title,
            "address": repair_request.address,
            "reference": str(repair_request.uuid),
            "payment": "Effectué en main propre au technicien."
        }
        return Response(data)


# ============================================================================
# VIEWSETS MANQUANTS
# ============================================================================


class ClientViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les clients."""

    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return Client.objects.select_related('user').all()


class TechnicianViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les techniciens."""

    queryset = Technician.objects.all()
    serializer_class = TechnicianSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return Technician.objects.select_related('user').all()

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def subscription_status(self, request):
        """Statut d'abonnement optimisé pour les techniciens."""
        user = request.user
        
        try:
            technician = get_technician_profile(user)
            if not technician:
                return Response({"error": "Profil technicien non trouvé"}, status=404)
            
            # Requête optimisée pour les abonnements
            active_subscription = TechnicianSubscription.objects.filter(
                technician=technician,
                is_active=True,
                end_date__gte=timezone.now()
            ).select_related('payment').first()
            
            if active_subscription:
                return Response({
                    'has_active_subscription': True,
                    'plan_name': active_subscription.plan_name,
                    'start_date': active_subscription.start_date,
                    'end_date': active_subscription.end_date,
                    'days_remaining': (active_subscription.end_date - timezone.now()).days
                })
            else:
                return Response({
                    'has_active_subscription': False,
                    'message': 'Aucun abonnement actif'
                })
                
        except Exception as e:
            logger.error(f"Erreur lors de la vérification du statut d'abonnement: {e}")
            return Response(
                {"error": "Erreur lors de la vérification du statut"},
                status=500
            )

    @action(detail=False, methods=["post"], permission_classes=[IsAuthenticated])
    def renew_subscription(self, request):
        """Renouvellement d'abonnement optimisé."""
        user = request.user
        
        try:
            technician = get_technician_profile(user)
            if not technician:
                return Response({"error": "Profil technicien non trouvé"}, status=404)
            
            # Créer une demande de paiement
            payment_request = SubscriptionPaymentRequest.objects.create(
                technician=technician,
                amount=Decimal("5000.00"),  # Prix fixe pour un an
                duration_months=12,
                payment_method="manual_validation",
                description="Abonnement gratuit d'un an pour le développement du Mali"
            )
            
            return Response({
                'success': True,
                'message': 'Demande de renouvellement créée',
                'payment_request_id': payment_request.id
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du renouvellement: {e}")
            return Response(
                {"error": "Erreur lors du renouvellement"},
                status=500
            )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def upload_photo(self, request, pk=None):
        """Upload de photo optimisé."""
        technician = self.get_object()
        
        if 'photo' not in request.FILES:
            return Response({"error": "Photo requise"}, status=400)
        
        try:
            photo = request.FILES['photo']
            # Validation du fichier
            if photo.size > 5 * 1024 * 1024:  # 5MB max
                return Response({"error": "Fichier trop volumineux"}, status=400)
            
            # Sauvegarde de la photo
            technician.user.photo = photo
            technician.user.save()
            
            return Response({
                'success': True,
                'message': 'Photo mise à jour avec succès'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de l'upload de photo: {e}")
            return Response(
                {"error": "Erreur lors de l'upload"},
                status=500
            )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def upload_kyc(self, request, pk=None):
        """Upload de documents KYC optimisé."""
        technician = self.get_object()
        
        if 'document' not in request.FILES:
            return Response({"error": "Document requis"}, status=400)
        
        try:
            document = request.FILES['document']
            # Validation du fichier
            if document.size > 10 * 1024 * 1024:  # 10MB max
                return Response({"error": "Fichier trop volumineux"}, status=400)
            
            # Sauvegarde du document
            technician.user.kyc_document = document
            technician.user.save()
            
            return Response({
                'success': True,
                'message': 'Document KYC mis à jour avec succès'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de l'upload KYC: {e}")
            return Response(
                {"error": "Erreur lors de l'upload"},
                status=500
            )

    @action(detail=True, methods=["get"], permission_classes=[IsAuthenticated])
    def download_receipts(self, request, pk=None):
        """Téléchargement de reçus optimisé."""
        technician = self.get_object()
        
        try:
            # Récupérer les paiements avec optimisations
            payments = Payment.objects.filter(
                recipient=technician.user,
                status='completed'
            ).select_related('request').order_by('-created_at')
            
            # Générer le fichier Excel
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reçus de paiement"
            
            # En-têtes
            headers = ['Date', 'Demande', 'Montant', 'Méthode', 'Statut']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Données
            for row, payment in enumerate(payments, 2):
                ws.cell(row=row, column=1, value=payment.created_at.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=2, value=payment.request.title if payment.request else 'N/A')
                ws.cell(row=row, column=3, value=float(payment.amount))
                ws.cell(row=row, column=4, value=payment.get_method_display())
                ws.cell(row=row, column=5, value=payment.get_status_display())
            
            # Sauvegarder le fichier
            filename = f"recus_technicien_{technician.user.username}_{timezone.now().strftime('%Y%m%d')}.xlsx"
            filepath = f"media/receipts/{filename}"
            
            import os
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)
            
            # Retourner le fichier
            from django.http import FileResponse
            response = FileResponse(open(filepath, 'rb'))
            response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"Erreur lors du téléchargement des reçus: {e}")
            return Response(
                {"error": "Erreur lors du téléchargement"},
                status=500
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def me(self, request):
        """Profil technicien optimisé."""
        user = request.user
        
        try:
            technician = get_technician_profile(user)
            if not technician:
                return Response({"error": "Profil technicien non trouvé"}, status=404)
            
            # Requêtes optimisées pour les statistiques
            stats = RepairRequest.objects.filter(
                technician=technician
            ).aggregate(
                total_requests=Count('id'),
                completed_requests=Count('id', filter=Q(status='completed')),
                total_earnings=Sum('final_price', filter=Q(status='completed')),
                avg_rating=Avg('review__rating', filter=Q(status='completed', review__isnull=False))
            )
            
            serializer = TechnicianSerializer(technician)
            data = serializer.data
            data.update({
                'total_requests': stats['total_requests'] or 0,
                'completed_requests': stats['completed_requests'] or 0,
                'total_earnings': float(stats['total_earnings'] or 0),
                'average_rating': round(stats['avg_rating'] or 0, 1)
            })
            
            return Response(data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération du profil: {e}")
            return Response(
                {"error": "Erreur lors de la récupération du profil"},
                status=500
            )


class RequestDocumentViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les documents de demande."""

    queryset = RequestDocument.objects.all()
    serializer_class = RequestDocumentSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return RequestDocument.objects.select_related('request', 'uploaded_by').all()


class ReviewViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les avis."""

    queryset = Review.objects.all()
    serializer_class = ReviewSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return Review.objects.select_related('request', 'client__user', 'technician__user').all()

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def received(self, request):
        """Avis reçus optimisés pour les techniciens."""
        user = request.user
        
        try:
            technician = get_technician_profile(user)
            if not technician:
                return Response({"error": "Profil technicien non trouvé"}, status=404)
            
            reviews = Review.objects.filter(
                technician=technician,
                is_visible=True,
                moderation_status='approved'
            ).select_related(
                'request', 'client__user'
            ).order_by('-created_at')
            
            # Pagination
            paginator = PageNumberPagination()
            paginator.page_size = 10
            paginated_reviews = paginator.paginate_queryset(reviews, request)
            
            serializer = ReviewSerializer(paginated_reviews, many=True)
            return paginator.get_paginated_response(serializer.data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des avis: {e}")
            return Response(
                {"error": "Erreur lors de la récupération des avis"},
                status=500
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def given(self, request):
        """Avis donnés par le client connecté."""
        user = request.user
        
        try:
            client = get_client_profile(user)
            if not client:
                return Response({"error": "Profil client non trouvé"}, status=404)
            
            reviews = Review.objects.filter(
                client=client,
                is_visible=True
            ).select_related(
                'request', 'technician__user'
            ).order_by('-created_at')
            
            # Pagination
            paginator = PageNumberPagination()
            paginator.page_size = 10
            paginated_reviews = paginator.paginate_queryset(reviews, request)
            
            serializer = ReviewSerializer(paginated_reviews, many=True)
            return paginator.get_paginated_response(serializer.data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des avis donnés: {e}")
            return Response(
                {"error": "Erreur lors de la récupération des avis"},
                status=500
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def statistics(self, request):
        """Statistiques d'avis optimisées."""
        user = request.user
        
        try:
            if user.user_type == 'technician':
                technician = get_technician_profile(user)
                if not technician:
                    return Response({"error": "Profil technicien non trouvé"}, status=404)
                
                # Requêtes optimisées avec annotations
                stats = Review.objects.filter(
                    technician=technician,
                    is_visible=True,
                    moderation_status='approved'
                ).aggregate(
                    total_reviews=Count('id'),
                    avg_rating=Avg('rating'),
                    avg_punctuality=Avg('punctuality_rating'),
                    avg_quality=Avg('quality_rating'),
                    avg_communication=Avg('communication_rating'),
                    avg_professionalism=Avg('professionalism_rating'),
                    avg_problem_solving=Avg('problem_solving_rating'),
                    avg_cleanliness=Avg('cleanliness_rating'),
                    avg_price_fairness=Avg('price_fairness_rating'),
                    would_recommend_count=Count('id', filter=Q(would_recommend=True)),
                    would_use_again_count=Count('id', filter=Q(would_use_again=True)),
                    would_recommend_friends_count=Count('id', filter=Q(would_recommend_to_friends=True)),
                    detailed_reviews_count=Count('id', filter=Q(
                        punctuality_rating__isnull=False,
                        quality_rating__isnull=False,
                        communication_rating__isnull=False
                    )),
                    verified_reviews_count=Count('id', filter=Q(is_verified_review=True))
                )
                
                return Response({
                    'total_reviews': stats['total_reviews'] or 0,
                    'average_rating': round(stats['avg_rating'] or 0, 1),
                    'average_punctuality': round(stats['avg_punctuality'] or 0, 1),
                    'average_quality': round(stats['avg_quality'] or 0, 1),
                    'average_communication': round(stats['avg_communication'] or 0, 1),
                    'average_professionalism': round(stats['avg_professionalism'] or 0, 1),
                    'average_problem_solving': round(stats['avg_problem_solving'] or 0, 1),
                    'average_cleanliness': round(stats['avg_cleanliness'] or 0, 1),
                    'average_price_fairness': round(stats['avg_price_fairness'] or 0, 1),
                    'recommendation_rate': round(
                        (stats['would_recommend_count'] / (stats['total_reviews'] or 1)) * 100, 1
                    ),
                    'reuse_rate': round(
                        (stats['would_use_again_count'] / (stats['total_reviews'] or 1)) * 100, 1
                    ),
                    'friend_recommendation_rate': round(
                        (stats['would_recommend_friends_count'] / (stats['total_reviews'] or 1)) * 100, 1
                    ),
                    'detailed_reviews_count': stats['detailed_reviews_count'] or 0,
                    'verified_reviews_count': stats['verified_reviews_count'] or 0
                })
            
            else:
                return Response({"error": "Statistiques disponibles uniquement pour les techniciens"}, status=403)
                
        except Exception as e:
            logger.error(f"Erreur lors du calcul des statistiques d'avis: {e}")
            return Response(
                {"error": "Erreur lors du calcul des statistiques"},
                status=500
            )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def update_review(self, request, pk=None):
        """Mise à jour d'avis optimisée."""
        review = self.get_object()
        user = request.user
        
        # Vérification des permissions
        if user.user_type == 'client' and review.client.user != user:
            return Response({"error": "Non autorisé"}, status=403)
        
        try:
            # Mise à jour des champs
            updateable_fields = [
                'rating', 'comment', 'punctuality_rating', 'quality_rating', 
                'communication_rating', 'professionalism_rating', 'problem_solving_rating',
                'cleanliness_rating', 'price_fairness_rating', 'would_recommend',
                'would_use_again', 'would_recommend_to_friends', 'positive_aspects',
                'improvement_suggestions', 'intervention_duration_minutes', 'was_urgent',
                'problem_complexity', 'parts_used', 'warranty_offered', 'warranty_duration_days',
                'tags'
            ]
            
            for field in updateable_fields:
                if field in request.data:
                    setattr(review, field, request.data[field])
            
            review.save()
            
            # Notification au technicien
            Notification.objects.create(
                recipient=review.technician.user,
                type=Notification.Type.REVIEW_RECEIVED,
                title="Avis mis à jour",
                message=f"Un client a mis à jour son avis sur votre travail",
                request=review.request
            )
            
            serializer = ReviewSerializer(review)
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour de l'avis: {e}")
            return Response(
                {"error": "Erreur lors de la mise à jour"},
                status=500
            )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def flag_review(self, request, pk=None):
        """Signaler un avis inapproprié."""
        review = self.get_object()
        user = request.user
        
        try:
            # Créer ou récupérer la modération
            moderation, created = ReviewModeration.objects.get_or_create(
                review=review,
                defaults={'status': 'flagged'}
            )
            
            if not created:
                moderation.status = 'flagged'
                moderation.save()
            
            # Ajouter l'utilisateur à la liste des signalements
            moderation.flagged_by_users.add(user)
            
            # Notification à l'admin
            AdminNotification.objects.create(
                title="Avis signalé",
                message=f"Avis #{review.id} signalé par {user.get_full_name()}",
                severity="warning",
                related_request=review.request,
                triggered_by=user
            )
            
            return Response({"message": "Avis signalé avec succès"})
            
        except Exception as e:
            logger.error(f"Erreur lors du signalement de l'avis: {e}")
            return Response(
                {"error": "Erreur lors du signalement"},
                status=500
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def analytics(self, request):
        """Analytics détaillés pour les techniciens."""
        user = request.user
        
        try:
            if user.user_type == 'technician':
                technician = get_technician_profile(user)
                if not technician:
                    return Response({"error": "Profil technicien non trouvé"}, status=404)
                
                # Récupérer ou créer les analytics
                analytics, created = ReviewAnalytics.objects.get_or_create(
                    technician=technician
                )
                
                # Recalculer les métriques
                analytics.calculate_all_metrics()
                
                serializer = ReviewAnalyticsSerializer(analytics)
                return Response(serializer.data)
            
            else:
                return Response({"error": "Analytics disponibles uniquement pour les techniciens"}, status=403)
                
        except Exception as e:
            logger.error(f"Erreur lors du calcul des analytics: {e}")
            return Response(
                {"error": "Erreur lors du calcul des analytics"},
                status=500
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def quality_metrics(self, request):
        """Métriques de qualité des avis."""
        user = request.user
        
        try:
            if user.user_type == 'technician':
                technician = get_technician_profile(user)
                if not technician:
                    return Response({"error": "Profil technicien non trouvé"}, status=404)
                
                reviews = Review.objects.filter(
                    technician=technician,
                    is_visible=True,
                    moderation_status='approved'
                )
                
                # Calculer les métriques de qualité
                quality_metrics = {
                    'total_reviews': reviews.count(),
                    'detailed_reviews_count': reviews.filter(
                        punctuality_rating__isnull=False,
                        quality_rating__isnull=False,
                        communication_rating__isnull=False
                    ).count(),
                    'verified_reviews_count': reviews.filter(is_verified_review=True).count(),
                    'avg_review_completeness': reviews.aggregate(
                        avg_completeness=Avg('review_quality_score')
                    )['avg_completeness'] or 0,
                    'avg_overall_score': reviews.aggregate(
                        avg_score=Avg('rating')
                    )['avg_score'] or 0,
                    'recent_trend': reviews.filter(
                        created_at__gte=timezone.now() - timedelta(days=30)
                    ).aggregate(
                        recent_avg=Avg('rating')
                    )['recent_avg'] or 0
                }
                
                return Response(quality_metrics)
            
            else:
                return Response({"error": "Métriques disponibles uniquement pour les techniciens"}, status=403)
                
        except Exception as e:
            logger.error(f"Erreur lors du calcul des métriques de qualité: {e}")
            return Response(
                {"error": "Erreur lors du calcul des métriques"},
                status=500
            )

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def popular_tags(self, request):
        """Tags populaires pour les avis."""
        user = request.user
        
        try:
            if user.user_type == 'technician':
                technician = get_technician_profile(user)
                if not technician:
                    return Response({"error": "Profil technicien non trouvé"}, status=404)
                
                reviews = Review.objects.filter(
                    technician=technician,
                    is_visible=True,
                    moderation_status='approved'
                )
                
                # Collecter tous les tags
                all_tags = []
                for review in reviews:
                    if review.tags:
                        all_tags.extend(review.tags)
                
                # Compter les occurrences
                tag_counts = {}
                for tag in all_tags:
                    tag_counts[tag] = tag_counts.get(tag, 0) + 1
                
                # Trier par popularité
                popular_tags = [
                    {'tag': tag, 'count': count}
                    for tag, count in sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)[:10]
                ]
                
                return Response({'popular_tags': popular_tags})
            
            else:
                return Response({"error": "Tags disponibles uniquement pour les techniciens"}, status=403)
                
        except Exception as e:
            logger.error(f"Erreur lors du calcul des tags populaires: {e}")
            return Response(
                {"error": "Erreur lors du calcul des tags"},
                status=500
            )


class PaymentViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les paiements."""

    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return Payment.objects.select_related('request', 'payer', 'recipient').all()

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def my_payments(self, request):
        """Paiements de l'utilisateur optimisés."""
        user = request.user
        
        try:
            # Récupérer les paiements selon le type d'utilisateur
            if user.user_type == 'technician':
                payments = Payment.objects.filter(
                    recipient=user
                ).select_related('request', 'payer')
            elif user.user_type == 'client':
                payments = Payment.objects.filter(
                    payer=user
                ).select_related('request', 'recipient')
            else:
                return Response({"error": "Type d'utilisateur non supporté"}, status=400)
            
            # Pagination
            paginator = PageNumberPagination()
            paginator.page_size = 20
            paginated_payments = paginator.paginate_queryset(payments, request)
            
            serializer = PaymentSerializer(paginated_payments, many=True)
            return paginator.get_paginated_response(serializer.data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des paiements: {e}")
            return Response(
                {"error": "Erreur lors de la récupération des paiements"},
                status=500
            )


class ConversationViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les conversations."""

    queryset = Conversation.objects.all()
    serializer_class = ConversationSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return Conversation.objects.select_related('request').prefetch_related('participants').all()

    @action(detail=True, methods=["post"], permission_classes=[IsAdminUser])
    def clear_messages(self, request, pk=None):
        """Efface tous les messages d'une conversation."""
        conversation = self.get_object()
        
        try:
            # Supprimer tous les messages
            Message.objects.filter(conversation=conversation).delete()
            
            return Response({
                'success': True,
                'message': 'Tous les messages ont été supprimés'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de la suppression des messages: {e}")
            return Response(
                {"error": "Erreur lors de la suppression"},
                status=500
            )


class MessageViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les messages."""

    queryset = Message.objects.all()
    serializer_class = MessageSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return Message.objects.select_related('conversation', 'sender').prefetch_related('attachments').all()

    @action(detail=False, methods=["post"], permission_classes=[IsAuthenticated])
    def mark_as_read(self, request):
        """Marque les messages comme lus."""
        user = request.user
        conversation_id = request.data.get('conversation_id')
        
        try:
            if conversation_id:
                # Marquer tous les messages d'une conversation comme lus
                Message.objects.filter(
                    conversation_id=conversation_id,
                    sender__in=Conversation.objects.get(id=conversation_id).participants.all()
                ).exclude(sender=user).update(
                    is_read=True,
                    read_at=timezone.now()
                )
            else:
                # Marquer tous les messages non lus de l'utilisateur
                Message.objects.filter(
                    conversation__participants=user
                ).exclude(sender=user).update(
                    is_read=True,
                    read_at=timezone.now()
                )
            
            return Response({
                'success': True,
                'message': 'Messages marqués comme lus'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du marquage des messages: {e}")
            return Response(
                {"error": "Erreur lors du marquage"},
                status=500
            )


class MessageAttachmentViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les pièces jointes des messages."""

    queryset = MessageAttachment.objects.all()
    serializer_class = MessageAttachmentSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return MessageAttachment.objects.select_related('message').all()


class NotificationViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les notifications."""

    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return Notification.objects.select_related('recipient', 'request').all()

    @action(detail=True, methods=["post"])
    def mark_as_read(self, request, pk=None):
        """Marque une notification comme lue."""
        notification = self.get_object()
        
        try:
            notification.mark_as_read()
            return Response({
                'success': True,
                'message': 'Notification marquée comme lue'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du marquage de la notification: {e}")
            return Response(
                {"error": "Erreur lors du marquage"},
                status=500
            )

    @action(detail=False, methods=["post"])
    def mark_all_as_read(self, request):
        """Marque toutes les notifications comme lues."""
        user = request.user
        
        try:
            Notification.objects.filter(
                recipient=user,
                is_read=False
            ).update(
                is_read=True,
                read_at=timezone.now()
            )
            
            return Response({
                'success': True,
                'message': 'Toutes les notifications marquées comme lues'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du marquage de toutes les notifications: {e}")
            return Response(
                {"error": "Erreur lors du marquage"},
                status=500
            )


class TechnicianLocationViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les localisations des techniciens."""

    queryset = TechnicianLocation.objects.all()
    serializer_class = TechnicianLocationSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return TechnicianLocation.objects.select_related('technician__user').all()


class SystemConfigurationViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer la configuration système."""

    queryset = SystemConfiguration.objects.all()
    serializer_class = SystemConfigurationSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        """Seuls les admins peuvent modifier la configuration."""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]


class TechnicianNearbyViewSet(viewsets.GenericViewSet):
    """ViewSet pour récupérer les techniciens proches avec géolocalisation optimisée."""
    permission_classes = [IsAuthenticated]
    serializer_class = TechnicianNearbySerializer
    pagination_class = PageNumberPagination

    def list(self, request):
        """Récupère les techniciens proches avec géolocalisation optimisée."""
        try:
            # Paramètres de géolocalisation
            latitude = request.query_params.get('latitude')
            longitude = request.query_params.get('longitude')
            specialty = request.query_params.get('specialty')
            max_distance = float(request.query_params.get('max_distance', 20))  # km
            
            # Base queryset optimisée
            queryset = Technician.objects.filter(
                is_available=True,
                is_verified=True
            ).select_related('user')
            
            # Filtrage par spécialité si spécifiée
            if specialty:
                queryset = queryset.filter(specialty=specialty)
            
            # Calcul des distances si géolocalisation fournie
            if latitude and longitude:
                try:
                    user_lat = float(latitude)
                    user_lon = float(longitude)
                    
                    technicians_with_distance = []
                    for technician in queryset:
                        if technician.current_latitude and technician.current_longitude:
                            distance = calculate_distance(
                                user_lat, user_lon,
                                technician.current_latitude, technician.current_longitude
                            )
                            if distance <= max_distance:
                                technicians_with_distance.append((technician, distance))
                    
                    # Tri par distance
                    technicians_with_distance.sort(key=lambda x: x[1])
                    technicians = [tech for tech, _ in technicians_with_distance]
                    
                except (ValueError, TypeError):
                    return Response(
                        {"error": "Coordonnées géographiques invalides"},
                        status=400
                    )
            else:
                # Sans géolocalisation, retourner tous les techniciens disponibles
                technicians = list(queryset)
            
            # Pagination
            paginator = PageNumberPagination()
            paginator.page_size = 20
            paginated_technicians = paginator.paginate_queryset(technicians, request)
            
            serializer = TechnicianNearbySerializer(paginated_technicians, many=True)
            return paginator.get_paginated_response(serializer.data)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des techniciens proches: {e}")
            return Response(
                {"error": "Erreur lors de la récupération des techniciens"},
                status=500
            )

    def calculate_haversine_distance(self, lat1, lon1, lat2, lon2):
        """Calcule la distance entre deux points géographiques."""
        from math import radians, cos, sin, asin, sqrt
        
        # Convertir en radians
        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        
        # Différences de coordonnées
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        # Formule de Haversine
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * asin(sqrt(a))
        
        # Rayon de la Terre en km
        r = 6371
        
        return c * r


class ClientLocationViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les localisations des clients."""
    queryset = ClientLocation.objects.all()
    serializer_class = ClientLocationSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return ClientLocation.objects.select_related('client__user').all()


class ReportViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les signalements."""
    queryset = Report.objects.all().order_by('-created_at')
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_permissions(self):
        """Seuls les admins peuvent modifier les signalements."""
        if self.action in ['update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        """Crée un signalement avec notification admin."""
        report = serializer.save()
        
        # Notification aux admins
        admin_users = User.objects.filter(is_staff=True, is_active=True)
        for admin in admin_users:
            AdminNotification.objects.create(
                title="Nouveau signalement",
                message=f"Signalement créé par {report.sender.username}",
                severity="warning",
                related_request=report.request,
                triggered_by=report.sender
            )

    def update(self, request, *args, **kwargs):
        """Met à jour un signalement avec validation."""
        report = self.get_object()
        
        try:
            # Mise à jour du statut
            new_status = request.data.get('status')
            if new_status and new_status in dict(Report.STATUS_CHOICES):
                report.status = new_status
                report.reviewed_at = timezone.now()
                report.reviewed_by = request.user
                report.save()
            
            return Response({
                'success': True,
                'message': f'Signalement mis à jour vers {new_status}'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour du signalement: {e}")
            return Response(
                {"error": "Erreur lors de la mise à jour"},
                status=500
            )


class AdminNotificationViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les notifications admin."""
    queryset = AdminNotification.objects.order_by('-created_at')
    serializer_class = AdminNotificationSerializer
    permission_classes = [permissions.IsAdminUser]
    pagination_class = PageNumberPagination

    def perform_update(self, serializer):
        """Met à jour une notification admin."""
        notification = serializer.save()
        
        # Log de la mise à jour
        logger.info(f"Notification admin mise à jour: {notification.id}")


class AuditLogListView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        logs = AuditLog.objects.all()
        # Filtres dynamiques
        event_type = request.GET.get('event_type')
        status = request.GET.get('status')
        user_email = request.GET.get('user_email')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if event_type:
            logs = logs.filter(event_type__icontains=event_type)
        if status:
            logs = logs.filter(status__icontains=status)
        if user_email:
            logs = logs.filter(user__email__icontains=user_email)
        if start_date:
            logs = logs.filter(timestamp__gte=start_date)
        if end_date:
            logs = logs.filter(timestamp__lte=end_date)
        logs = logs.order_by('-timestamp')[:100]
        serializer = AuditLogSerializer(logs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="export", permission_classes=[permissions.IsAdminUser])
    def export(self, request):
        logs = AuditLog.objects.all()
        # Filtres dynamiques
        event_type = request.GET.get('event_type')
        status = request.GET.get('status')
        user_email = request.GET.get('user_email')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if event_type:
            logs = logs.filter(event_type__icontains=event_type)
        if status:
            logs = logs.filter(status__icontains=status)
        if user_email:
            logs = logs.filter(user__email__icontains=user_email)
        if start_date:
            logs = logs.filter(timestamp__gte=start_date)
        if end_date:
            logs = logs.filter(timestamp__lte=end_date)
        logs = logs.order_by('-timestamp')
        export_format = request.GET.get('format', 'csv')
        fields = [
            'timestamp', 'user__email', 'event_type', 'status', 'ip_address', 'geo_country', 'geo_city', 'user_agent', 'risk_score', 'metadata'
        ]
        if export_format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Audit Log"
            ws.append(['Date', 'Utilisateur', 'Type', 'Statut', 'IP', 'Pays', 'Ville', 'User Agent', 'Risque', 'Détails'])
            for log in logs:
                ws.append([
                    log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    getattr(log.user, 'email', ''),
                    log.event_type,
                    log.status,
                    log.ip_address,
                    log.geo_country,
                    log.geo_city,
                    log.user_agent,
                    log.risk_score,
                    str(log.metadata) if log.metadata else ''
                ])
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="audit_log.xlsx"'
            return response
        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="audit_log.csv"'
            writer = csv.writer(response)
            writer.writerow(['Date', 'Utilisateur', 'Type', 'Statut', 'IP', 'Pays', 'Ville', 'User Agent', 'Risque', 'Détails'])
            for log in logs:
                writer.writerow([
                    log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    getattr(log.user, 'email', ''),
                    log.event_type,
                    log.status,
                    log.ip_address,
                    log.geo_country,
                    log.geo_city,
                    log.user_agent,
                    log.risk_score,
                    str(log.metadata) if log.metadata else ''
                ])
            return response


class ChatConversationViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les conversations de chat optimisées."""
    
    serializer_class = ChatConversationSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        user = self.request.user
        return ChatConversation.objects.filter(
            Q(client=user) | Q(technician=user)
        ).select_related(
            'client', 'technician', 'request'
        ).prefetch_related('messages').order_by('-last_message_at')

    @action(detail=False, methods=['post'], url_path='get_or_create', url_name='get_or_create')
    def get_or_create_conversation(self, request):
        import traceback
        try:
            client_id = request.data.get('client_id')
            technician_id = request.data.get('technician_id')
            if not client_id or not technician_id:
                return Response({'error': 'client_id et technician_id sont requis'}, status=400)
            client = User.objects.get(id=client_id)
            technician = User.objects.get(id=technician_id)
            from .models import ChatConversation
            conversation, created = ChatConversation.objects.get_or_create(
                client=client,
                technician=technician
            )
            serializer = self.get_serializer(conversation)
            return Response(serializer.data, status=201 if created else 200)
        except User.DoesNotExist:
            logger.error('Utilisateur non trouvé lors de la création de conversation', exc_info=True)
            return Response({'error': 'Utilisateur non trouvé'}, status=404)
        except Exception as e:
            logger.error(f'Erreur lors de la création de conversation: {e}\n{traceback.format_exc()}', exc_info=True)
            return Response({'error': str(e)}, status=500)

    @action(detail=True, methods=['post'])
    def mark_as_read(self, request, pk=None):
        """Marque une conversation comme lue."""
        conversation = self.get_object()
        user = request.user
        
        try:
            conversation.mark_all_as_read_for_user(user)
            return Response({
                'success': True,
                'message': 'Conversation marquée comme lue'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du marquage de la conversation: {e}")
            return Response(
                {"error": "Erreur lors du marquage"},
                status=500
            )

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        """Archive une conversation."""
        conversation = self.get_object()
        
        try:
            conversation.is_active = False
            conversation.save()
            
            return Response({
                'success': True,
                'message': 'Conversation archivée'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de l'archivage: {e}")
            return Response(
                {"error": "Erreur lors de l'archivage"},
                status=500
            )


class ChatMessageViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les messages de chat optimisés."""
    
    serializer_class = ChatMessageSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        user = self.request.user
        return ChatMessage.objects.filter(
            conversation__in=ChatConversation.objects.filter(
                Q(client=user) | Q(technician=user)
            )
        ).select_related('conversation', 'sender').prefetch_related('attachments')

    def perform_create(self, serializer):
        """Crée un message avec optimisations."""
        message = serializer.save()
        
        try:
            # Mettre à jour la conversation
            conversation = message.conversation
            conversation.last_message_at = timezone.now()
            conversation.save()
            
            # Notification à l'autre utilisateur
            other_user = conversation.client if message.sender == conversation.technician else conversation.technician
            Notification.objects.create(
                recipient=other_user,
                type=Notification.Type.MESSAGE_RECEIVED,
                title="Nouveau message",
                message=f"Nouveau message de {message.sender.get_full_name()}",
                request=conversation.request
            )
            
        except Exception as e:
            logger.error(f"Erreur lors de la création du message: {e}")

    @action(detail=False, methods=['get'])
    def conversation_messages(self, request):
        """Récupère les messages d'une conversation optimisée."""
        conversation_id = request.query_params.get('conversation_id')
        
        try:
            conversation = ChatConversation.objects.get(id=conversation_id)
            
            # Vérifier les permissions
            user = request.user
            if user not in [conversation.client, conversation.technician]:
                return Response({"error": "Accès non autorisé"}, status=403)
            
            messages = ChatMessage.objects.filter(
                conversation=conversation
            ).select_related('sender').prefetch_related('attachments').order_by('created_at')
            
            # Pagination
            paginator = PageNumberPagination()
            paginator.page_size = 50
            paginated_messages = paginator.paginate_queryset(messages, request)
            
            serializer = ChatMessageSerializer(paginated_messages, many=True)
            return paginator.get_paginated_response(serializer.data)
            
        except ChatConversation.DoesNotExist:
            return Response({"error": "Conversation non trouvée"}, status=404)
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des messages: {e}")
            return Response(
                {"error": "Erreur lors de la récupération des messages"},
                status=500
            )

    @action(detail=True, methods=['post'])
    def mark_as_read(self, request, pk=None):
        """Marque un message comme lu."""
        message = self.get_object()
        
        try:
            message.mark_as_read()
            return Response({
                'success': True,
                'message': 'Message marqué comme lu'
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du marquage du message: {e}")
            return Response(
                {"error": "Erreur lors du marquage"},
                status=500
            )

    @action(detail=False, methods=['post'])
    def send_location(self, request):
        """Envoie un message de localisation."""
        conversation_id = request.data.get('conversation_id')
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        
        try:
            conversation = ChatConversation.objects.get(id=conversation_id)
            
            # Vérifier les permissions
            user = request.user
            if user not in [conversation.client, conversation.technician]:
                return Response({"error": "Accès non autorisé"}, status=403)
            
            # Créer le message de localisation
            message = ChatMessage.objects.create(
                conversation=conversation,
                sender=user,
                content="Localisation partagée",
                message_type=ChatMessage.MessageType.LOCATION,
                latitude=latitude,
                longitude=longitude
            )
            
            serializer = ChatMessageSerializer(message)
            return Response(serializer.data)
            
        except ChatConversation.DoesNotExist:
            return Response({"error": "Conversation non trouvée"}, status=404)
        except Exception as e:
            logger.error(f"Erreur lors de l'envoi de localisation: {e}")
            return Response(
                {"error": "Erreur lors de l'envoi de localisation"},
                status=500
            )


class ChatMessageAttachmentViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les pièces jointes de chat optimisées."""
    
    serializer_class = ChatMessageAttachmentSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        """Optimise les requêtes avec select_related."""
        return ChatMessageAttachment.objects.select_related('message__conversation').all()

    def perform_create(self, serializer):
        """Crée une pièce jointe avec optimisations."""
        attachment = serializer.save()
        
        try:
            # Mettre à jour la taille du fichier
            attachment.file_size = attachment.file.size
            attachment.save()
            
        except Exception as e:
            logger.error(f"Erreur lors de la création de la pièce jointe: {e}")


class ChatGetOrCreateConversationView(APIView):
    """Vue optimisée pour récupérer ou créer une conversation."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Récupère ou crée une conversation optimisée."""
        user = request.user
        other_user_id = request.data.get('other_user_id')
        request_id = request.data.get('request_id')
        
        try:
            other_user = User.objects.get(id=other_user_id)
            
            # Chercher une conversation existante avec optimisations
            conversation = ChatConversation.objects.filter(
                Q(client=user, technician=other_user) |
                Q(client=other_user, technician=user)
            ).select_related('request').first()
            
            if not conversation:
                # Créer une nouvelle conversation
                conversation = ChatConversation.objects.create(
                    client=user if user.user_type == 'client' else other_user,
                    technician=user if user.user_type == 'technician' else other_user,
                    request_id=request_id,
                    is_active=True
                )
            
            serializer = ChatConversationSerializer(conversation)
            return Response(serializer.data)
            
        except User.DoesNotExist:
            return Response({"error": "Utilisateur non trouvé"}, status=404)
        except Exception as e:
            import traceback
            logger.error(f"Erreur lors de la création de conversation: {e}\n{traceback.format_exc()}")
            return Response(
                {"error": str(e), "trace": traceback.format_exc()},
                status=500
            )


class ChatStatsView(APIView):
    """Vue pour récupérer les statistiques de communication."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Récupère les statistiques de communication pour une demande."""
        request_id = request.query_params.get('request_id')
        
        try:
            repair_request = RepairRequest.objects.get(id=request_id)
            
            # Vérifier les permissions
            user = request.user
            if user not in [repair_request.client, repair_request.technician]:
                return Response({"error": "Accès non autorisé"}, status=403)
            
            # Récupérer la conversation
            conversation = ChatConversation.objects.filter(
                request=repair_request
            ).first()
            
            if not conversation:
                return Response({
                    "total_messages": 0,
                    "unread_messages": 0,
                    "last_activity": None,
                    "response_time_avg": 0
                })
            
            # Calculer les statistiques
            total_messages = ChatMessage.objects.filter(conversation=conversation).count()
            unread_messages = conversation.unread_count_for_user(user)
            last_activity = conversation.last_message_at
            
            # Calculer le temps de réponse moyen
            messages = ChatMessage.objects.filter(
                conversation=conversation
            ).exclude(sender=user).order_by('created_at')
            
            response_times = []
            for i, msg in enumerate(messages):
                if i > 0:
                    prev_msg = messages[i-1]
                    if prev_msg.sender == user:  # Le message précédent était de l'utilisateur
                        time_diff = (msg.created_at - prev_msg.created_at).total_seconds() / 60
                        response_times.append(time_diff)
            
            response_time_avg = sum(response_times) / len(response_times) if response_times else 0
            
            return Response({
                "total_messages": total_messages,
                "unread_messages": unread_messages,
                "last_activity": last_activity.isoformat() if last_activity else None,
                "response_time_avg": round(response_time_avg, 1)
            })
            
        except RepairRequest.DoesNotExist:
            return Response({"error": "Demande non trouvée"}, status=404)
        except Exception as e:
            logger.error(f"Erreur lors du calcul des statistiques: {e}")
            return Response(
                {"error": "Erreur lors du calcul des statistiques"},
                status=500
            )


class SendLocationView(APIView):
    """Vue pour envoyer un message de localisation."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Envoie un message de localisation."""
        conversation_id = request.data.get('conversation_id')
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        
        try:
            conversation = ChatConversation.objects.get(id=conversation_id)
            
            # Vérifier les permissions
            user = request.user
            if user not in [conversation.client, conversation.technician]:
                return Response({"error": "Accès non autorisé"}, status=403)
            
            # Créer le message de localisation
            message = ChatMessage.objects.create(
                conversation=conversation,
                sender=user,
                content="📍 Ma position actuelle",
                message_type='location',
                latitude=latitude,
                longitude=longitude
            )
            
            # Mettre à jour la conversation
            conversation.last_message_at = timezone.now()
            conversation.save()
            
            # Notification à l'autre utilisateur
            other_user = conversation.client if user == conversation.technician else conversation.technician
            Notification.objects.create(
                recipient=other_user,
                type=Notification.Type.LOCATION_SHARED,
                title="Position partagée",
                message=f"{user.get_full_name()} a partagé sa position",
                request=conversation.request
            )
            
            serializer = ChatMessageSerializer(message)
            return Response(serializer.data)
            
        except ChatConversation.DoesNotExist:
            return Response({"error": "Conversation non trouvée"}, status=404)
        except Exception as e:
            logger.error(f"Erreur lors de l'envoi de la localisation: {e}")
            return Response(
                {"error": "Erreur lors de l'envoi de la localisation"},
                status=500
            )


class VoiceMessageView(APIView):
    """Vue pour envoyer un message vocal."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Envoie un message vocal."""
        conversation_id = request.data.get('conversation')
        audio_file = request.FILES.get('file')
        
        try:
            conversation = ChatConversation.objects.get(id=conversation_id)
            
            # Vérifier les permissions
            user = request.user
            if user not in [conversation.client, conversation.technician]:
                return Response({"error": "Accès non autorisé"}, status=403)
            
            # Créer le message vocal
            message = ChatMessage.objects.create(
                conversation=conversation,
                sender=user,
                content="🎤 Message vocal",
                message_type='voice'
            )
            
            # Sauvegarder le fichier audio
            if audio_file:
                # Créer le dossier si nécessaire
                import os
                audio_dir = os.path.join(settings.MEDIA_ROOT, 'voice_messages')
                os.makedirs(audio_dir, exist_ok=True)
                
                # Générer un nom de fichier unique
                import uuid
                filename = f"voice_{uuid.uuid4()}.wav"
                file_path = os.path.join(audio_dir, filename)
                
                with open(file_path, 'wb+') as destination:
                    for chunk in audio_file.chunks():
                        destination.write(chunk)
                
                # Créer l'attachement
                ChatMessageAttachment.objects.create(
                    message=message,
                    file=os.path.join('voice_messages', filename),
                    file_name=audio_file.name,
                    content_type=audio_file.content_type
                )
            
            # Mettre à jour la conversation
            conversation.last_message_at = timezone.now()
            conversation.save()
            
            # Notification à l'autre utilisateur
            other_user = conversation.client if user == conversation.technician else conversation.technician
            Notification.objects.create(
                recipient=other_user,
                type=Notification.Type.MESSAGE_RECEIVED,
                title="Nouveau message vocal",
                message=f"Message vocal de {user.get_full_name()}",
                request=conversation.request
            )
            
            serializer = ChatMessageSerializer(message)
            return Response(serializer.data)
            
        except ChatConversation.DoesNotExist:
            return Response({"error": "Conversation non trouvée"}, status=404)
        except Exception as e:
            logger.error(f"Erreur lors de l'envoi du message vocal: {e}")
            return Response(
                {"error": "Erreur lors de l'envoi du message vocal"},
                status=500
            )


class CommunicationDashboardView(APIView):
    """Vue pour le tableau de bord de communication."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Récupère les données du tableau de bord de communication."""
        user = request.user
        
        try:
            # Récupérer les conversations de l'utilisateur
            conversations = ChatConversation.objects.filter(
                Q(client=user) | Q(technician=user)
            ).select_related('client', 'technician', 'request').prefetch_related('messages')
            
            # Statistiques générales
            total_conversations = conversations.count()
            unread_conversations = sum(1 for conv in conversations if conv.unread_count_for_user(user) > 0)
            
            # Messages récents
            recent_messages = ChatMessage.objects.filter(
                conversation__in=conversations
            ).select_related('conversation', 'sender').order_by('-created_at')[:10]
            
            # Demandes actives avec communication
            active_requests = RepairRequest.objects.filter(
                Q(client=user) | Q(technician=user),
                status__in=['accepted', 'in_progress']
            ).select_related('client', 'technician').prefetch_related('chat_conversation')
            
            return Response({
                "total_conversations": total_conversations,
                "unread_conversations": unread_conversations,
                "recent_messages": ChatMessageSerializer(recent_messages, many=True).data,
                "active_requests": RepairRequestSerializer(active_requests, many=True).data
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du chargement du tableau de bord: {e}")
            return Response(
                {"error": "Erreur lors du chargement du tableau de bord"},
                status=500
            )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def techniciens_proches(request):
    """API optimisée pour trouver les techniciens les plus proches avec géolocalisation précise."""
    try:
        # Paramètres de géolocalisation
        lat = request.query_params.get('lat')
        lng = request.query_params.get('lng')
        
        # Paramètres de filtrage
        specialty = request.query_params.get('specialty')
        min_experience_level = request.query_params.get('min_experience_level')
        min_rating = request.query_params.get('min_rating')
        urgence = request.query_params.get('urgence', 'normal')
        max_distance = float(request.query_params.get('max_distance', 30))  # 30km par défaut
        limit = int(request.query_params.get('limit', 20))
        
        # Validation des coordonnées
        if not lat or not lng:
            return Response(
                {"error": "Les coordonnées lat et lng sont requises"},
                status=400
            )
        
        try:
            user_lat = float(lat)
            user_lng = float(lng)
        except ValueError:
            return Response(
                {"error": "Coordonnées géographiques invalides"},
                status=400
            )
        
        # Base queryset optimisée
        queryset = Technician.objects.filter(
            is_available=True,
            is_verified=True
        ).select_related('user', 'location')
        
        # Filtrage par spécialité
        if specialty:
            queryset = queryset.filter(specialty=specialty)
        
        # Filtrage par niveau d'expérience
        if min_experience_level:
            queryset = queryset.filter(experience_level__gte=min_experience_level)
        
        # Filtrage par note minimale
        if min_rating:
            try:
                min_rating_float = float(min_rating)
                queryset = queryset.filter(average_rating__gte=min_rating_float)
            except ValueError:
                pass
        
        # Récupération des techniciens avec calcul de distance
        technicians_with_distance = []
        
        for technician in queryset:
            # Utiliser la localisation du technicien si disponible
            tech_lat = None
            tech_lng = None
            
            if hasattr(technician, 'location') and technician.location:
                tech_lat = technician.location.latitude
                tech_lng = technician.location.longitude
            elif technician.current_latitude and technician.current_longitude:
                tech_lat = technician.current_latitude
                tech_lng = technician.current_longitude
            
            if tech_lat and tech_lng:
                # Calcul de distance avec formule de Haversine
                distance = calculate_distance(user_lat, user_lng, tech_lat, tech_lng)
                
                # Vérifier le rayon d'intervention du technicien
                if distance <= max_distance and distance <= technician.service_radius_km:
                    # Calcul du temps d'arrivée estimé
                    eta_minutes = calculate_eta(distance, urgence)
                    
                    technicians_with_distance.append({
                        'technician': technician,
                        'distance': round(distance, 2),
                        'eta_minutes': eta_minutes,
                        'location_quality': get_location_quality(technician),
                        'is_urgent_available': technician.is_available_urgent and urgence == 'urgent'
                    })
        
        # Tri par distance et disponibilité urgente
        technicians_with_distance.sort(key=lambda x: (x['is_urgent_available'], x['distance']))
        
        # Limiter le nombre de résultats
        technicians_with_distance = technicians_with_distance[:limit]
        
        # Préparer la réponse
        technicians_data = []
        for item in technicians_with_distance:
            tech = item['technician']
            serializer = TechnicianNearbySerializer(tech)
            tech_data = serializer.data
            tech_data.update({
                'distance': item['distance'],
                'eta_minutes': item['eta_minutes'],
                'location_quality': item['location_quality'],
                'is_urgent_available': item['is_urgent_available'],
                'current_location': {
                    'latitude': tech.location.latitude if hasattr(tech, 'location') and tech.location else tech.current_latitude,
                    'longitude': tech.location.longitude if hasattr(tech, 'location') and tech.location else tech.current_longitude,
                    'accuracy': tech.location.accuracy if hasattr(tech, 'location') and tech.location else None,
                    'is_moving': tech.location.is_moving if hasattr(tech, 'location') and tech.location else False,
                    'last_update': tech.location.updated_at.isoformat() if hasattr(tech, 'location') and tech.location else None
                }
            })
            technicians_data.append(tech_data)
        
        # Statistiques de recherche
        search_stats = {
            'total_found': len(technicians_data),
            'search_radius_km': max_distance,
            'user_location': {
                'latitude': user_lat,
                'longitude': user_lng
            },
            'filters_applied': {
                'specialty': specialty,
                'min_experience_level': min_experience_level,
                'min_rating': min_rating,
                'urgence': urgence
            }
        }
        
        return Response({
            'technicians': technicians_data,
            'search_stats': search_stats,
            'message': f"Trouvé {len(technicians_data)} technicien(s) dans un rayon de {max_distance}km"
        })
        
    except Exception as e:
        logger.error(f"Erreur lors de la recherche de techniciens proches: {e}")
        return Response(
            {"error": "Erreur lors de la recherche de techniciens"},
            status=500
        )


def calculate_eta(distance_km, urgence='normal'):
    """Calcule le temps d'arrivée estimé en minutes."""
    # Vitesses moyennes selon le niveau d'urgence
    speeds = {
        'normal': 25,      # 25 km/h en ville
        'urgent': 35,      # 35 km/h pour urgence
        'sos': 45          # 45 km/h pour SOS
    }
    
    speed = speeds.get(urgence, speeds['normal'])
    eta_minutes = int((distance_km / speed) * 60)
    
    # Ajouter un délai de préparation
    preparation_time = {
        'normal': 15,
        'urgent': 10,
        'sos': 5
    }
    
    eta_minutes += preparation_time.get(urgence, 15)
    
    return eta_minutes


def get_location_quality(technician):
    """Détermine la qualité de la localisation du technicien."""
    if hasattr(technician, 'location') and technician.location:
        accuracy = technician.location.accuracy
        if accuracy is None:
            return 'unknown'
        elif accuracy <= 10:
            return 'excellent'
        elif accuracy <= 50:
            return 'good'
        elif accuracy <= 100:
            return 'fair'
        else:
            return 'poor'
    else:
        return 'unknown'


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def techniciens_proches_avances(request):
    """API avancée pour la recherche de techniciens avec filtres multiples."""
    try:
        # Paramètres de géolocalisation
        lat = request.query_params.get('lat')
        lng = request.query_params.get('lng')
        
        # Paramètres de filtrage avancés
        specialties = request.query_params.getlist('specialties[]')
        experience_levels = request.query_params.getlist('experience_levels[]')
        rating_range = request.query_params.get('rating_range', '3.0-5.0')
        price_range = request.query_params.get('price_range', '0-10000')
        availability_time = request.query_params.get('availability_time', 'now')
        max_distance = float(request.query_params.get('max_distance', 30))
        limit = int(request.query_params.get('limit', 20))
        
        # Validation des coordonnées
        if not lat or not lng:
            return Response(
                {"error": "Les coordonnées lat et lng sont requises"},
                status=400
            )
        
        try:
            user_lat = float(lat)
            user_lng = float(lng)
        except ValueError:
            return Response(
                {"error": "Coordonnées géographiques invalides"},
                status=400
            )
        
        # Base queryset
        queryset = Technician.objects.filter(
            is_available=True,
            is_verified=True
        ).select_related('user', 'location')
        
        # Filtrage par spécialités multiples
        if specialties:
            queryset = queryset.filter(specialty__in=specialties)
        
        # Filtrage par niveaux d'expérience
        if experience_levels:
            queryset = queryset.filter(experience_level__in=experience_levels)
        
        # Filtrage par note
        try:
            min_rating, max_rating = map(float, rating_range.split('-'))
            queryset = queryset.filter(average_rating__gte=min_rating, average_rating__lte=max_rating)
        except (ValueError, AttributeError):
            pass
        
        # Filtrage par prix
        try:
            min_price, max_price = map(float, price_range.split('-'))
            queryset = queryset.filter(hourly_rate__gte=min_price, hourly_rate__lte=max_price)
        except (ValueError, AttributeError):
            pass
        
        # Récupération et calcul des distances
        technicians_with_details = []
        
        for technician in queryset:
            tech_lat = None
            tech_lng = None
            
            if hasattr(technician, 'location') and technician.location:
                tech_lat = technician.location.latitude
                tech_lng = technician.location.longitude
            elif technician.current_latitude and technician.current_longitude:
                tech_lat = technician.current_latitude
                tech_lng = technician.current_longitude
            
            if tech_lat and tech_lng:
                distance = calculate_distance(user_lat, user_lng, tech_lat, tech_lng)
                
                if distance <= max_distance and distance <= technician.service_radius_km:
                    # Calcul des métriques avancées
                    eta_normal = calculate_eta(distance, 'normal')
                    eta_urgent = calculate_eta(distance, 'urgent')
                    
                    technicians_with_details.append({
                        'technician': technician,
                        'distance': round(distance, 2),
                        'eta_normal': eta_normal,
                        'eta_urgent': eta_urgent,
                        'location_quality': get_location_quality(technician),
                        'availability_score': calculate_availability_score(technician),
                        'reliability_score': calculate_reliability_score(technician)
                    })
        
        # Tri par score de disponibilité et distance
        technicians_with_details.sort(
            key=lambda x: (x['availability_score'], x['reliability_score'], x['distance']),
            reverse=True
        )
        
        # Limiter les résultats
        technicians_with_details = technicians_with_details[:limit]
        
        # Préparer la réponse
        technicians_data = []
        for item in technicians_with_details:
            tech = item['technician']
            serializer = TechnicianNearbySerializer(tech)
            tech_data = serializer.data
            tech_data.update({
                'distance': item['distance'],
                'eta_normal': item['eta_normal'],
                'eta_urgent': item['eta_urgent'],
                'location_quality': item['location_quality'],
                'availability_score': item['availability_score'],
                'reliability_score': item['reliability_score'],
                'current_location': {
                    'latitude': tech.location.latitude if hasattr(tech, 'location') and tech.location else tech.current_latitude,
                    'longitude': tech.location.longitude if hasattr(tech, 'location') and tech.location else tech.current_longitude,
                    'accuracy': tech.location.accuracy if hasattr(tech, 'location') and tech.location else None,
                    'is_moving': tech.location.is_moving if hasattr(tech, 'location') and tech.location else False,
                    'last_update': tech.location.updated_at.isoformat() if hasattr(tech, 'location') and tech.location else None
                }
            })
            technicians_data.append(tech_data)
        
        return Response({
            'technicians': technicians_data,
            'search_metadata': {
                'total_found': len(technicians_data),
                'search_radius_km': max_distance,
                'user_location': {'latitude': user_lat, 'longitude': user_lng},
                'filters_applied': {
                    'specialties': specialties,
                    'experience_levels': experience_levels,
                    'rating_range': rating_range,
                    'price_range': price_range,
                    'availability_time': availability_time
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Erreur lors de la recherche avancée de techniciens: {e}")
        return Response(
            {"error": "Erreur lors de la recherche avancée de techniciens"},
            status=500
        )


def calculate_availability_score(technician):
    """Calcule un score de disponibilité du technicien."""
    score = 0
    
    # Base score pour disponibilité
    if technician.is_available:
        score += 30
    
    # Bonus pour disponibilité urgente
    if technician.is_available_urgent:
        score += 20
    
    # Bonus pour temps de réponse
    if technician.response_time_minutes <= 15:
        score += 25
    elif technician.response_time_minutes <= 30:
        score += 15
    elif technician.response_time_minutes <= 60:
        score += 10
    
    # Bonus pour expérience
    if technician.years_experience >= 10:
        score += 15
    elif technician.years_experience >= 5:
        score += 10
    elif technician.years_experience >= 2:
        score += 5
    
    return min(score, 100)


def calculate_reliability_score(technician):
    """Calcule un score de fiabilité du technicien."""
    score = 0
    
    # Base score pour vérification
    if technician.is_verified:
        score += 25
    
    # Score basé sur la note moyenne
    if technician.average_rating >= 4.5:
        score += 30
    elif technician.average_rating >= 4.0:
        score += 25
    elif technician.average_rating >= 3.5:
        score += 20
    elif technician.average_rating >= 3.0:
        score += 15
    
    # Bonus pour niveau d'expérience
    if technician.experience_level == 'expert':
        score += 20
    elif technician.experience_level == 'senior':
        score += 15
    elif technician.experience_level == 'intermediate':
        score += 10
    
    # Bonus pour badge
    if technician.badge_level == 'platinum':
        score += 15
    elif technician.badge_level == 'gold':
        score += 10
    elif technician.badge_level == 'silver':
        score += 5
    
    return min(score, 100)


class StatisticsViewSet(viewsets.ViewSet):
    """Vues pour les statistiques avancées."""
    
    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def global_statistics(self, request):
        """Récupère les statistiques globales de la plateforme."""
        user = request.user
        
        if user.user_type != "admin":
            return Response(
                {"error": "Accès non autorisé"}, status=403
            )
        
        try:
            # Vérifier le cache d'abord
            cache_key = f"global_stats_{timezone.now().strftime('%Y-%m-%d')}"
            cached_data = StatisticsCache.get_valid_cache(cache_key)
            
            if cached_data:
                return Response(cached_data)
            
            # Calculer les nouvelles statistiques
            stats, created = GlobalStatistics.objects.get_or_create(
                id=1,  # Une seule instance de statistiques globales
                defaults={}
            )
            
            # Recalculer toutes les métriques
            stats.calculate_all_metrics()
            
            # Préparer la réponse
            response_data = {
                'overview': {
                    'total_users': stats.total_users,
                    'total_clients': stats.total_clients,
                    'total_technicians': stats.total_technicians,
                    'total_admins': stats.total_admins,
                    'active_users_30d': stats.active_users_30d,
                    'new_users_30d': stats.new_users_30d,
                    'total_requests': stats.total_requests,
                    'completed_requests': stats.completed_requests,
                    'total_revenue': float(stats.total_revenue),
                    'platform_fees': float(stats.platform_fees),
                    'avg_rating': stats.avg_rating,
                    'satisfaction_rate': stats.satisfaction_rate
                },
                'requests': {
                    'total': stats.total_requests,
                    'pending': stats.pending_requests,
                    'in_progress': stats.in_progress_requests,
                    'completed': stats.completed_requests,
                    'cancelled': stats.cancelled_requests,
                    'urgent': stats.urgent_requests,
                    'success_rate': stats.success_rate,
                    'avg_response_time_hours': stats.avg_response_time_hours,
                    'avg_completion_time_hours': stats.avg_completion_time_hours
                },
                'financial': {
                    'total_revenue': float(stats.total_revenue),
                    'total_payouts': float(stats.total_payouts),
                    'platform_fees': float(stats.platform_fees),
                    'avg_request_value': float(stats.avg_request_value),
                    'payment_methods': stats.payment_methods,
                    'payment_success_rate': stats.payment_success_rate
                },
                'satisfaction': {
                    'total_reviews': stats.total_reviews,
                    'avg_rating': stats.avg_rating,
                    'satisfaction_rate': stats.satisfaction_rate,
                    'recommendation_rate': stats.recommendation_rate
                },
                'technicians': {
                    'total': stats.total_technicians,
                    'verified': stats.verified_technicians,
                    'available': stats.available_technicians,
                    'avg_rating': stats.avg_technician_rating,
                    'top_technicians': stats.top_technicians
                },
                'specialties': {
                    'distribution': stats.specialty_distribution,
                    'top_specialties': stats.top_specialties
                },
                'security': {
                    'total_logins': stats.total_logins,
                    'failed_logins': stats.failed_logins,
                    'security_alerts': stats.security_alerts,
                    'login_success_rate': stats.login_success_rate
                },
                'geography': {
                    'top_cities': stats.top_cities,
                    'service_areas': stats.service_areas
                },
                'trends': {
                    'daily': stats.daily_stats,
                    'weekly': stats.weekly_stats,
                    'monthly': stats.monthly_stats
                },
                'advanced': {
                    'conversion_rate': stats.conversion_rate,
                    'retention_rate': stats.retention_rate,
                    'churn_rate': stats.churn_rate
                },
                'calculation_info': {
                    'last_calculation': stats.last_calculation.isoformat(),
                    'calculation_duration': stats.calculation_duration
                }
            }
            
            # Mettre en cache pour 1 heure
            StatisticsCache.get_or_create_cache(cache_key, response_data, expires_in_hours=1)
            
            return Response(response_data)
            
        except Exception as e:
            logger.error(f"Erreur lors du calcul des statistiques globales: {e}")
            return Response(
                {"error": "Erreur lors du calcul des statistiques"},
                status=500
            )
    
    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def real_time_stats(self, request):
        """Statistiques en temps réel (sans cache)."""
        user = request.user
        
        if user.user_type != "admin":
            return Response(
                {"error": "Accès non autorisé"}, status=403
            )
        
        try:
            from django.utils import timezone
            from datetime import timedelta
            
            now = timezone.now()
            last_24h = now - timedelta(hours=24)
            last_7d = now - timedelta(days=7)
            
            # Statistiques en temps réel
            real_time_data = {
                'current_time': now.isoformat(),
                'last_24h': {
                    'new_requests': RepairRequest.objects.filter(created_at__gte=last_24h).count(),
                    'completed_requests': RepairRequest.objects.filter(
                        status='completed', completed_at__gte=last_24h
                    ).count(),
                    'new_users': User.objects.filter(created_at__gte=last_24h).count(),
                    'new_reviews': Review.objects.filter(created_at__gte=last_24h).count(),
                    'revenue': float(Payment.objects.filter(
                        status='completed',
                        payment_type='client_payment',
                        created_at__gte=last_24h
                    ).aggregate(total=Sum('amount'))['total'] or 0)
                },
                'last_7d': {
                    'new_requests': RepairRequest.objects.filter(created_at__gte=last_7d).count(),
                    'completed_requests': RepairRequest.objects.filter(
                        status='completed', completed_at__gte=last_7d
                    ).count(),
                    'new_users': User.objects.filter(created_at__gte=last_7d).count(),
                    'new_reviews': Review.objects.filter(created_at__gte=last_7d).count(),
                    'revenue': float(Payment.objects.filter(
                        status='completed',
                        payment_type='client_payment',
                        created_at__gte=last_7d
                    ).aggregate(total=Sum('amount'))['total'] or 0)
                },
                'active_sessions': {
                    'online_users': User.objects.filter(last_login__gte=last_24h).count(),
                    'active_technicians': Technician.objects.filter(
                        is_available=True, is_verified=True
                    ).count()
                }
            }
            
            return Response(real_time_data)
            
        except Exception as e:
            logger.error(f"Erreur lors du calcul des statistiques temps réel: {e}")
            return Response(
                {"error": "Erreur lors du calcul des statistiques temps réel"},
                status=500
            )
    
    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def export_statistics(self, request):
        """Export des statistiques en différents formats."""
        user = request.user
        
        if user.user_type != "admin":
            return Response(
                {"error": "Accès non autorisé"}, status=403
            )
        
        try:
            export_type = request.query_params.get('type', 'excel')
            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to')
            
            # Créer une demande d'export
            export_request = StatisticsExport.objects.create(
                export_type=export_type,
                requested_by=user,
                export_config={
                    'date_from': date_from,
                    'date_to': date_to,
                    'filters': dict(request.query_params)
                }
            )
            
            # Simuler le traitement (en production, ceci serait géré par une tâche asynchrone)
            import os
            from django.conf import settings
            
            # Créer le fichier d'export
            export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
            os.makedirs(export_dir, exist_ok=True)
            
            filename = f"statistics_export_{export_request.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
            
            if export_type == 'excel':
                filename += '.xlsx'
                # Ici, vous utiliseriez openpyxl ou xlsxwriter pour créer le fichier Excel
                file_path = os.path.join(export_dir, filename)
                with open(file_path, 'w') as f:
                    f.write("Statistiques exportées\n")
                
            elif export_type == 'csv':
                filename += '.csv'
                file_path = os.path.join(export_dir, filename)
                with open(file_path, 'w') as f:
                    f.write("Date,Utilisateurs,Demandes,Revenus\n")
                
            elif export_type == 'json':
                filename += '.json'
                file_path = os.path.join(export_dir, filename)
                import json
                with open(file_path, 'w') as f:
                    json.dump({'statistics': 'data'}, f)
            
            # Mettre à jour l'export
            file_size = os.path.getsize(file_path)
            export_request.file_path = file_path
            export_request.file_size_bytes = file_size
            export_request.status = 'completed'
            export_request.save()
            
            return Response({
                'export_id': export_request.id,
                'status': 'completed',
                'download_url': f'/media/exports/{filename}',
                'file_size': file_size
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export des statistiques: {e}")
            return Response(
                {"error": "Erreur lors de l'export des statistiques"},
                status=500
            )
    
    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def dashboard_widgets(self, request):
        """Récupère les widgets de tableau de bord personnalisés."""
        user = request.user
        
        try:
            # Trouver le tableau de bord approprié pour l'utilisateur
            dashboard_type = 'admin' if user.user_type == 'admin' else user.user_type
            dashboard = StatisticsDashboard.objects.filter(
                dashboard_type=dashboard_type,
                is_active=True
            ).first()
            
            if not dashboard:
                return Response({
                    'dashboard': None,
                    'widgets': []
                })
            
            widgets_data = []
            for widget in dashboard.widgets.filter(is_visible=True).order_by('position_y', 'position_x'):
                widgets_data.append({
                    'id': widget.id,
                    'name': widget.name,
                    'type': widget.widget_type,
                    'data_source': widget.data_source,
                    'config': widget.config,
                    'position': {
                        'x': widget.position_x,
                        'y': widget.position_y,
                        'width': widget.width,
                        'height': widget.height
                    },
                    'refresh_interval': widget.refresh_interval_seconds,
                    'last_update': widget.last_update.isoformat() if widget.last_update else None
                })
            
            return Response({
                'dashboard': {
                    'id': dashboard.id,
                    'name': dashboard.name,
                    'description': dashboard.description,
                    'type': dashboard.dashboard_type,
                    'layout_config': dashboard.layout_config,
                    'refresh_interval': dashboard.refresh_interval_seconds
                },
                'widgets': widgets_data
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des widgets: {e}")
            return Response(
                {"error": "Erreur lors de la récupération des widgets"},
                status=500
            )
    
    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def alerts(self, request):
        """Récupère les alertes de statistiques actives."""
        user = request.user
        
        if user.user_type != "admin":
            return Response(
                {"error": "Accès non autorisé"}, status=403
            )
        
        try:
            alerts = StatisticsAlert.objects.filter(
                is_active=True
            ).order_by('-triggered_at')[:50]
            
            alerts_data = []
            for alert in alerts:
                alerts_data.append({
                    'id': alert.id,
                    'type': alert.alert_type,
                    'title': alert.title,
                    'message': alert.message,
                    'severity': alert.severity,
                    'triggered_at': alert.triggered_at.isoformat(),
                    'resolved_at': alert.resolved_at.isoformat() if alert.resolved_at else None,
                    'triggered_by': alert.triggered_by.get_full_name() if alert.triggered_by else None,
                    'resolved_by': alert.resolved_by.get_full_name() if alert.resolved_by else None
                })
            
            return Response({
                'alerts': alerts_data,
                'total_active': StatisticsAlert.objects.filter(is_active=True).count()
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des alertes: {e}")
            return Response(
                {"error": "Erreur lors de la récupération des alertes"},
                status=500
            )


class InitiateSubscriptionPaymentView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = CinetPayInitiationSerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            user = request.user
            amount = data["amount"]
            phone = data.get("phone")
            description = data["description"]
            # Appel logique métier
            result, transaction_id = init_cinetpay_payment(
                amount=amount,
                phone=phone,
                name=user.get_full_name(),
                description=description,
                user=user
            )
            if result.get("success"):
                return Response(result, status=201)
            return Response(result, status=400)
        return Response(serializer.errors, status=400)
