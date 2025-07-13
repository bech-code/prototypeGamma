import io
import datetime
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .models import RepairRequest, Payment, Review
from users.models import User
from django.db.models import Count, Sum, Avg, Q

@api_view(['GET'])
@permission_classes([IsAdminUser])
def export_statistics_excel(request):
    """Export des statistiques en Excel."""
    if not request.user.is_staff:
        return Response({"error": "Accès non autorisé"}, status=403)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from django.db.models import Q, Count, Sum, Avg
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        # Calculer les statistiques
        now = timezone.now()
        
        # Utilisateurs actifs (correction de la requête)
        active_users_30d = User.objects.filter(
            Q(client_profile__repair_requests__created_at__gte=now - timedelta(days=30)) |
            Q(technician_depannage__repair_requests__created_at__gte=now - timedelta(days=30))
        ).distinct().count()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistiques"

        # Statistiques utilisateurs
        total_users = User.objects.count()
        total_clients = User.objects.filter(user_type='client').count()
        total_technicians = User.objects.filter(user_type='technician').count()
        total_admins = User.objects.filter(user_type='admin').count()
        ws.append(["Vue d'ensemble"])
        ws.append(["Total utilisateurs", total_users])
        ws.append(["Clients", total_clients])
        ws.append(["Techniciens", total_technicians])
        ws.append(["Admins", total_admins])
        ws.append(["Utilisateurs actifs (30j)", active_users_30d])
        ws.append([])

        # Statistiques demandes
        ws.append(["Demandes"])
        total_requests = RepairRequest.objects.count()
        completed_requests = RepairRequest.objects.filter(status="completed").count()
        pending_requests = RepairRequest.objects.filter(status="pending").count()
        in_progress_requests = RepairRequest.objects.filter(status="in_progress").count()
        cancelled_requests = RepairRequest.objects.filter(status="cancelled").count()
        ws.append(["Total demandes", total_requests])
        ws.append(["Terminées", completed_requests])
        ws.append(["En attente", pending_requests])
        ws.append(["En cours", in_progress_requests])
        ws.append(["Annulées", cancelled_requests])
        ws.append([])

        # Statistiques financières
        ws.append(["Finances"])
        total_revenue = Payment.objects.filter(status='completed', payment_type='client_payment').aggregate(total=Sum('amount'))['total'] or 0
        total_payouts = Payment.objects.filter(status='completed', payment_type='technician_payout').aggregate(total=Sum('amount'))['total'] or 0
        ws.append(["Revenus totaux (XOF)", float(total_revenue)])
        ws.append(["Paiements techniciens (XOF)", float(total_payouts)])
        ws.append(["Frais plateforme (XOF)", float(total_revenue) - float(total_payouts)])
        ws.append([])

        # Statistiques satisfaction
        ws.append(["Satisfaction"])
        total_reviews = Review.objects.count()
        avg_rating = Review.objects.aggregate(avg=Avg('rating'))['avg'] or 0
        ws.append(["Nombre d'avis", total_reviews])
        ws.append(["Note moyenne", round(avg_rating, 2)])
        ws.append([])

        # Statistiques par spécialité
        ws.append(["Demandes par spécialité"])
        specialty_stats = (
            RepairRequest.objects.values('specialty_needed')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        ws.append(["Spécialité", "Nombre de demandes"])
        for s in specialty_stats:
            ws.append([s['specialty_needed'], s['count']])
        ws.append([])

        # Statistiques par ville
        ws.append(["Demandes par ville"])
        city_stats = (
            RepairRequest.objects.values('city')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        ws.append(["Ville", "Nombre de demandes"])
        for c in city_stats:
            ws.append([c['city'], c['count']])
        ws.append([])

        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # Générer le fichier Excel en mémoire
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"statistiques_depanneteliman_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response 
    except ImportError:
        return Response({"error": "Les dépendances nécessaires pour l'export Excel sont manquantes."}, status=500) 